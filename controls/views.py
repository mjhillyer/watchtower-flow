import functools
import logging
import operator
import pathlib
import random
import shutil
import tempfile
import ast
from collections import defaultdict
from datetime import datetime
from itertools import groupby
from pathlib import PurePath
from urllib.parse import quote, urlunparse
from uuid import uuid4
import re
import rtyaml
import trestle.oscal.component as trestlecomponent
import trestle.oscal.ssp as trestlessp
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.db.models.functions import Lower
from django.http import Http404, HttpResponse, HttpResponseRedirect, HttpResponseForbidden, JsonResponse, \
    HttpResponseNotAllowed
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.utils import timezone
from django.views import View
from django.views.generic import ListView
from django.urls import reverse
from jsonschema import validate
from jsonschema.exceptions import SchemaError, ValidationError as SchemaValidationError
from urllib.parse import quote

from api.siteapp.serializers.tags import SimpleTagSerializer
from guidedmodules.models import Task, Module, AppVersion, AppSource
from guidedmodules.app_loading import ModuleDefinitionError
from siteapp.model_mixins.tags import TagView
from simple_history.utils import update_change_reason

from siteapp.models import Project, Organization, Folder, Portfolio, Tag, User, Role, Party, Appointment, Request, Proposal
from siteapp.settings import GOVREADY_URL
from siteapp.utils.views_helper import project_context, start_app, get_compliance_apps_catalog, \
    get_compliance_apps_catalog_for_user, get_compliance_apps_catalog_for_user
from system_settings.models import SystemSettings
from .forms import ElementEditForm, ElementEditAccessManagementForm
from .forms import ImportOSCALComponentForm, SystemAssessmentResultForm
from .forms import StatementPoamForm, PoamForm, ElementForm, DeploymentForm, StatementEditForm, ImportSystemForm, ImportProjectForm
from .models import *
from .utilities import *
from integrations.models import Integration

logging.basicConfig()
import structlog
from structlog import get_logger
from structlog.stdlib import LoggerFactory
structlog.configure(logger_factory=LoggerFactory())
structlog.configure(processors=[structlog.processors.JSONRenderer()])
logger = get_logger()

from django.template.defaulttags import register

def index(request):
    """Index page for controls"""

    # Get catalog
    catalog = Catalog()
    cg_flat = catalog.get_flattened_controls_all_as_dict()
    control_groups = catalog.get_groups()
    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "control_groups": control_groups
    }
    return render(request, "controls/index.html", context)

def catalogs(request):
    """Index page for catalogs"""

    context = {
        "catalogs": Catalogs(),
    }
    return render(request, "controls/index-catalogs.html", context)

def catalog(request, catalog_key, system_id=None):
    """Index page for controls"""

    if system_id is None:
        system = None
    else:
        system = System.objects.get(pk=system_id)

    # Get catalog
    catalog = Catalog(catalog_key)
    cg_flat = catalog.get_flattened_controls_all_as_dict()
    control_groups = catalog.get_groups()
    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "system": system,
        "control_groups": control_groups,
    }
    return render(request, "controls/index.html", context)

def group(request, catalog_key, g_id):
    """Temporary index page for catalog control group"""

    # Get catalog
    catalog = Catalog(catalog_key)
    cg_flat = catalog.get_flattened_controls_all_as_dict()
    control_groups = catalog.get_groups()
    group = None
    # Get group/family of controls
    for g in control_groups:
        if g['id'].lower() == g_id:
            group = g
            break

    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "control_groups": control_groups,
        "group": group,
    }
    return render(request, "controls/index-group.html", context)

def control(request, catalog_key, cl_id):
    """Control detail view"""
    control_id = cl_id
    cl_id = oscalize_control_id(cl_id)
    catalog_key = oscalize_catalog_key(catalog_key)

    # Get catalog
    catalog = Catalog(catalog_key)
    cg_flat = catalog.get_flattened_controls_all_as_dict()
    # Handle properly formatted control id that does not exist
    if cl_id.lower() not in cg_flat:
        return render(request, "controls/detail.html", {"control_id": control_id, "catalog": catalog,"control": {}})
    # Prepare links
    links = []
    for link in cg_flat[cl_id.lower()]['guidance_links']:
        link['href_split'] = link['href'].split("/")
        if len(link['href_split']) == 6:
            link['catalog'] = link['href_split'][3].replace("_"," ")
        else:
            link['catalog'] = None
        links.append(link)

    # Get and return the control
    context = {
        "catalog": catalog,
        "control": cg_flat[cl_id.lower()],
        "links": links,
    }
    return render(request, "controls/detail.html", context)

@functools.lru_cache()
def controls_selected_aspen(request, system_id):
    """Display System's selected controls view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # if not request.user.has_perm('view_system', system):
    #     raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    # system_events = get_integrations_system_events(request, system_id)

    # Fix menu data for old vertical menu
    project.system.root_element.name = system_summary['name']
    project.root_task.title_override = system_summary['name']

    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        impl_smts = system.root_element.statements_consumed.all()

        # sort controls
        controls = list(controls)
        controls.sort(key=lambda control: control.get_flattened_oscal_control_as_dict()['sort_id'])

        # Determine if a legacy statement exists for the control
        impl_smts_legacy = Statement.objects.filter(consumer_element=system.root_element, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_LEGACY.name)
        impl_smts_legacy_dict = {}
        for legacy_smt in impl_smts_legacy:
            impl_smts_legacy_dict[legacy_smt.sid] = legacy_smt

        # Get count of componentes (e.g., producer_elements) associated with a control
        impl_smts_cmpts_count = {}
        ikeys = system.smts_control_implementation_as_dict.keys()
        for c in controls:
            impl_smts_cmpts_count[c.oscal_ctl_id] = 0
            if c.oscal_ctl_id in ikeys:
                impl_smts_cmpts_count[c.oscal_ctl_id] = len(set([s.producer_element for s in system.smts_control_implementation_as_dict[c.oscal_ctl_id]['control_impl_smts']]))

        # Get list of catalog objects
        catalog_list = Catalogs().list_catalogs()
        # Remove the 3 nist catalogs that are hard-coded already in template
        external_catalogs = [catalog for catalog in catalog_list if catalog.catalog_key not in ['NIST_SP-800-53_rev4', 'NIST_SP-800-53_rev5', 'NIST_SP-800-171_rev1', 'CMMC_ver1' ]]

        # Return the controls
        context = {
            "system": system,
            "system_summary": system_summary,
            "project": project,
            "controls": controls,
            "external_catalogs": external_catalogs,
            "impl_smts_cmpts_count": impl_smts_cmpts_count,
            "impl_smts_legacy_dict": impl_smts_legacy_dict,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "display_urls": project_context(project)
        }
        return render(request, "systems/controls_selected_aspen.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@functools.lru_cache()
def controls_selected(request, system_id):
    """Display System's selected controls view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        impl_smts = system.root_element.statements_consumed.all()

        # sort controls
        controls = list(controls)
        controls.sort(key=lambda control: control.get_flattened_oscal_control_as_dict()['sort_id'])

        # Determine if a legacy statement exists for the control
        impl_smts_legacy = Statement.objects.filter(consumer_element=system.root_element, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_LEGACY.name)
        impl_smts_legacy_dict = {}
        for legacy_smt in impl_smts_legacy:
            impl_smts_legacy_dict[legacy_smt.sid] = legacy_smt

        # Get count of componentes (e.g., producer_elements) associated with a control
        impl_smts_cmpts_count = {}
        ikeys = system.smts_control_implementation_as_dict.keys()
        for c in controls:
            impl_smts_cmpts_count[c.oscal_ctl_id] = 0
            if c.oscal_ctl_id in ikeys:
                impl_smts_cmpts_count[c.oscal_ctl_id] = len(set([s.producer_element for s in system.smts_control_implementation_as_dict[c.oscal_ctl_id]['control_impl_smts']]))

        # Get list of catalog objects
        catalog_list = Catalogs().list_catalogs()
        # Remove the 3 nist catalogs that are hard-coded already in template
        external_catalogs = [catalog for catalog in catalog_list if catalog.catalog_key not in ['NIST_SP-800-53_rev4', 'NIST_SP-800-53_rev5', 'NIST_SP-800-171_rev1', 'CMMC_ver1' ]]

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "controls": controls,
            "external_catalogs": external_catalogs,
            "impl_smts_cmpts_count": impl_smts_cmpts_count,
            "impl_smts_legacy_dict": impl_smts_legacy_dict,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "display_urls": project_context(project)
        }
        return render(request, "systems/controls_selected.html", context)
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def system_controls_add(request, system_id):
    """Add a selected control to a system (e.g., selected controls)"""

    # Get control values from request.POST
    catalog_key = request.POST['catalog_key'].replace(" ","_") # Make sure catalog key has underscores instead of spaces
    control_id = request.POST['control_id']

    system = System.objects.get(id=system_id)
    controls = system.root_element.controls.all()

    # If the catalog key and control id combination returns a result than don't add to controls selected
    if controls.filter(Q(oscal_catalog_key=catalog_key)).filter(Q(oscal_ctl_id=control_id)):
        messages.add_message(request, messages.WARNING, f"Control {control_id.upper()} in catalog {catalog_key} is already in selected controls!")
        # Log result
        logger.warning(
                event="change_system add_selected_control",
                object={"object": "control", "id": control_id, "catalog": catalog_key},
                user={"id": request.user.id, "username": request.user.username}
                )
        return redirect(reverse('controls_selected', args=[system_id]))

    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('change_system', system):

        # Add ElementControl to system
        system.add_control(catalog_key, control_id)

        # Create message for user
        messages.add_message(request, messages.INFO, f"Control {control_id.upper()} added to selected controls.")

    else:
        # User does not have permission
        # Log result
        logger.info(
                event="change_system add_selected_control permission_denied",
                object={"object": "control", "id": control_id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        messages.add_message(request, messages.INFO, f"You do not have permission to edit the system.")

    response = redirect(reverse('controls_selected', args=[system_id]))
    return response

@login_required
def system_control_remove(request, system_id, element_control_id):
    """Remove a selected control from a system and delete/hide the related statements"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('change_system', system):

        # Retrieve ElementControl
        ec = ElementControl.objects.get(id=element_control_id)

        # Delete the control implementation statements associated with this component
        system.remove_control(element_control_id)
        # result = element.statements_produced.filter(consumer_element=system.root_element).delete()
        messages.add_message(request, messages.INFO, f"Removed control '{ec.oscal_ctl_id}' from system.")

        # Log result
        logger.info(
                event="change_system remove_selected_control",
                object={"object": "control", "id": element_control_id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        # messages.add_message(request, messages.INFO, f"Removed control '{element_control_id}' from system.")

    else:
        # User does not have permission
        # Log result
        logger.info(
                event="change_system remove_selected_control permission_denied",
                object={"object": "control", "id": element_control_id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        messages.add_message(request, messages.INFO, f"You do not have permission to edit the system.")

    response = redirect(reverse('controls_selected', args=[system_id]))
    return response

@login_required
def system_proposal_remove(request, system_id, proposal_id):
    """Remove a proposal from a system"""
    
        # Retrieve identified System
    system = System.objects.get(id=system_id)
    proposal = Proposal.objects.get(id=proposal_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('change_system', system):
        system.remove_proposals([proposal.id])
        if proposal.req != None:
            proposal.req.status = 'Closed'
            proposal.req.save()
        Proposal.objects.filter(id=proposal_id).delete()
        messages.add_message(request, messages.INFO, f"Removed proposal '{proposal.requested_element.name}' from system.")
        # Log result
        logger.info(
                event="change_system remove_selected_proposal",
                object={"object": "proposal", "id": proposal_id},
                user={"id": request.user.id, "username": request.user.username}
                )
    else:
        # User does not have permission
        # Log result
        logger.info(
                event="change_system remove_selected_proposal permission_denied",
                object={"object": "proposal", "id": proposal_id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        messages.add_message(request, messages.INFO, f"You do not have permission to edit the system.")
    response = redirect(reverse('components_selected', args=[system_id]))
    return response

@functools.lru_cache()
def controls_updated(request, system_id):
    """Display System's statements by updated date in reverse chronological order"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        impl_smts = system.root_element.statements_consumed.all()

        impl_smts_count = {}
        ikeys = system.smts_control_implementation_as_dict.keys()
        for c in controls:
            impl_smts_count[c.oscal_ctl_id] = 0
            if c.oscal_ctl_id in ikeys:
                impl_smts_count[c.oscal_ctl_id] = len(
                    system.smts_control_implementation_as_dict[c.oscal_ctl_id]['control_impl_smts'])

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "controls": controls,
            "impl_smts_count": impl_smts_count,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
        }
        return render(request, "systems/controls_updated.html", context)
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def edit_element(request, element_id):
    """
      Edit Element information as long as the name does not clash with a different element name
    Args:
        request ([HttpRequest]): The network request
    component_id ([int|str]): The id of the component
    Returns:
        [JsonResponse]: Either a ok status or an error
    """

    # The original element(component)
    ele_instance = get_object_or_404(Element, id=element_id)

    if request.method == 'POST':
        new_name = request.POST.get("name", "").strip() or None

        # Check if the new component name is already in use and if the new name is different from the current name
        if Element.objects.filter(name__iexact=new_name).exists() and new_name != ele_instance.name:
            return JsonResponse({"status": "err", "message": "Name already in use"})

        form = ElementEditForm(request.POST or None, instance=ele_instance)
        if form.is_valid():
            logger.info(
                event="edit_element",
                object={"object": "element", "id": form.instance.id, "name": form.instance.name},
                user={"id": request.user.id, "username": request.user.username}
            )
            form.save()
            return JsonResponse({"status": "ok"})
        else:
            errors = form.errors.get_json_data(escape_html=False)
            msg_list = [f"{e.title()} - {errors[e][0]['message']}" for e in errors.keys()]
            return JsonResponse({"status": "err", "message": "Please fix the following problems:<br>"+"<br>".join(msg_list)})

class SelectedComponentsList(ListView):
    """
    Display System's selected components view
    """
    model = Element
    template_name = 'systems/components_selected.html'
    context_object_name = 'system_elements'
    ordering = ['name']
    paginate_by = 25

    def get_queryset(self):
        """
        Return the systems producer elements.
        """
        system = System.objects.get(id=self.kwargs['system_id'])
        return [element for element in system.producer_elements if element.element_type != "system"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve identified System
        
        system = System.objects.get(id=self.kwargs['system_id'])

        system_proposals = []
        system_proposal_elements = []
        for proposal in system.proposals.all():
            system_proposals.append(proposal)
            system_proposal_elements.append(proposal.requested_element)
        # Retrieve related selected controls if user has permission on system
        if self.request.user.has_perm('view_system', system):
            # Retrieve primary system Project
            # Temporarily assume only one project and get first project
            project = system.projects.first()
            context['project'] = project
            context['system'] = system
            context['system_proposals'] = system_proposals
            context['system_proposal_elements'] =  system_proposal_elements
            context['elements'] = Element.objects.all().exclude(element_type='system')
            context["display_urls"] = project_context(project)
            return context
        else:
            # User does not have permission to this system
            raise Http404

class SelectedComponentsListAspen(ListView):
    """
    Display System's selected components view
    """
    model = Element
    template_name = 'systems/components_selected_aspen.html'
    context_object_name = 'system_elements'
    ordering = ['name']
    paginate_by = 25

    def get_queryset(self):
        """
        Return the systems producer elements.
        """
        system = System.objects.get(id=self.kwargs['system_id'])
        return [element for element in system.producer_elements if element.element_type != "system"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve identified System

        system = System.objects.get(id=self.kwargs['system_id'])
        system_summary = get_integrations_system_info(self.request, self.kwargs['system_id'])

        system_proposals = []
        system_proposal_elements = []
        for proposal in system.proposals.all():
            system_proposals.append(proposal)
            system_proposal_elements.append(proposal.requested_element)
        # Retrieve related selected controls if user has permission on system
        if self.request.user.has_perm('view_system', system):
            # Retrieve primary system Project
            # Temporarily assume only one project and get first project
            project = system.projects.first()
            context['project'] = project
            context['system_summary'] = system_summary
            context['system'] = system
            context['system_proposals'] = system_proposals
            context['system_proposal_elements'] =  system_proposal_elements
            context['elements'] = Element.objects.all().exclude(element_type='system')
            context["display_urls"] = project_context(project)
            return context
        else:
            # User does not have permission to this system
            raise Http404


@login_required
def component_library(request):
    """Display the library of components"""
    owned_elements_id = []

    for element in Element.objects.all().exclude(element_type='system').distinct():
        if element.is_owner(request.user):
            owned_elements_id.append(element.id)
    
    owned_elements_list = Element.objects.filter(id__in=owned_elements_id)
    
    query = request.GET.get('search')
    # Setting a breakpoint in the code.
    if query:
        try:
            if request.GET.get('owner'):
                # Search by owner
                element_list = Element.objects.filter(id__in=owned_elements_id)
                element_list = element_list.filter(Q(name__icontains=query) | Q(tags__label__icontains=query)).exclude(element_type='system').distinct()
            else:
                element_list = Element.objects.filter(Q(name__icontains=query) | Q(tags__label__icontains=query)).exclude(element_type='system').distinct()

        except:
            logger.info(f"Ah, you are not using Postgres for your Database!")
            element_list = Element.objects.filter(Q(name__icontains=query) | Q(tags__label__icontains=query)).exclude(element_type='system').distinct()
    else:
        element_list = Element.objects.all().exclude(element_type='system').distinct()

    # Natural sorting on name
    element_list = natsorted(element_list, key=lambda x: x.name)

    # Remove private elements for which user does not have permission
    element_list_private_removed = [element for element in element_list if (element.private==False or 'view_element' in get_user_perms(request.user, element))]

    # Pagination
    ele_paginator = Paginator(element_list_private_removed, 15)
    owned_ele_paginator = Paginator(owned_elements_list, 15)
    page_number = request.GET.get('page')

    try:
        page_obj = ele_paginator.page(page_number)
        owned_page_obj = owned_ele_paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = ele_paginator.page(1)
        owned_page_obj = owned_ele_paginator.page(1)
    except EmptyPage:
        page_obj = ele_paginator.page(ele_paginator.num_pages)
        owned_page_obj = owned_ele_paginator.page(owned_ele_paginator.num_pages)
    
    context = {
        "page_obj": page_obj,
        "owned_page_obj": owned_page_obj,
        "start": False,
        "import_form": ImportOSCALComponentForm(),
        "total_comps": Element.objects.exclude(element_type='system').count(),
    }

    return render(request, "components/component_library.html", context)

def diff_components_prettyHtml(smt1, smt2):
    """Generate a diff of two statements of type `control_implementation`"""
    dmp = dmp_module.diff_match_patch()
    val1 = ""
    val2 = ""
    if hasattr(smt1, 'body'):
        val1 = smt1.body
    if hasattr(smt2, 'body'):
        val2 = smt2.body

    diff = dmp.diff_main(val1, val2)
    if len(diff) == 1:
        return "Statement is identical."
    return dmp.diff_prettyHtml(diff)

def compare_components(request):
    """
    Compare submitted components
    """

    checks = json.loads(request.POST.get('hiddenChecks'))
    compare_list = list(checks.values())
    if len(compare_list) <= 1:
        # add messages
        messages.add_message(request, messages.WARNING, f"Not enough components were selected to compare!")
        return HttpResponseRedirect("/controls/components")
    else:
        ele_q = Element.objects.filter(pk__in=compare_list).exclude(element_type='system').distinct()
        # Maintain sort order of compare_list otherwise Django will order ascending
        element_list = sorted(ele_q, key=lambda x: compare_list.index(str(x.id)))
        compare_prime, element_list = element_list[0], element_list[
                                                       1:]  # The first component selected will be compared against the rest
        compare_prime_smts = compare_prime.statements(StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name)
    difference_tuples = []
    for component in element_list:
        differences = []
        # compare each component's statements to prime
        cmt_smts = component.statements(StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name)
        if cmt_smts.exists():
            # TODO: Need to create a tuple with smt id to return appropriate
            for smt in cmt_smts:
                smt_prime = compare_prime_smts.filter(sid=smt.sid).filter(pid=smt.pid).filter(sid_class=smt.sid_class).first()
                # If it is not a statement found in both components then we just add styling
                if smt_prime:
                    diff = diff_components_prettyHtml(smt_prime, smt)
                else:
                    diff = f"<span><ins style='background:#e6ffe6;'>{smt.body}</ins><span>"
                differences.append(diff)
        difference_tuples.extend(zip([component.id] * len(cmt_smts),[component.name] * len(cmt_smts), cmt_smts, differences))
    if request.method == 'POST':
        context = {
            "element_list": element_list,
            "compare_prime": compare_prime,
            "prime_smts": compare_prime_smts,
            "secondary_smts": cmt_smts,
            "differences": difference_tuples,
            "compare_list": compare_list
        }
        return render(request, "components/compare_components.html", context)

@login_required
def import_records(request):
    """Display the records of component imports"""

    import_records = ImportRecord.objects.all()
    import_components = {}

    for import_record in import_records:
        import_components[import_record] = Element.objects.filter(import_record=import_record)

    context = {
        "import_components": import_components,
    }

    return render(request, "components/import_records.html", context)

@login_required
def import_record_details(request, import_record_id):
    """Display the records of component imports"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    component_statements = import_record.get_components_statements()
    statements_prototype = Statement.objects.filter(statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
                                                    import_record=import_record)

    context = {
        "import_record": import_record,
        "component_statements": component_statements,
        "statements_prototype": statements_prototype,
    }
    return render(request, "components/import_record_details.html", context)

@login_required
def confirm_import_record_delete(request, import_record_id):
    """Delete the components and statements imported from a particular import record"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    component_statements = import_record.get_components_statements()
    component_count = len(component_statements)
    statement_count = 0
    for component in component_statements:
        statement_count += component_statements[component].count()
    statements_prototype = Statement.objects.filter(statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
                                                    import_record=import_record)
    statement_prototype_count = len(statements_prototype)
    projects = import_record.import_record_projects.all()
    project_count = len(projects)
    elements = import_record.import_record_elements.all()
    element_count = len(elements)
    context = {
        "import_record": import_record,
        "component_count": component_count,
        "statement_count": statement_count,
        "statement_prototype_count": statement_prototype_count,
        "project_count": project_count,
        "element_count": element_count
    }
    return render(request, "components/confirm_import_record_delete.html", context)

@login_required
def import_record_delete(request, import_record_id):
    """Delete the components and statements imported from a particular import record"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    import_created = import_record.created
    import_record.delete()

    messages.add_message(request, messages.INFO, f"Deleted import: {import_created}")

    response = redirect('/controls/components')
    return response

class SystemSecurityPlanSerializer(object):

    def __init__(self, system, impl_smts):
        self.system = system
        self.impl_smts = impl_smts

class OSCALSystemSecurityPlanSerializer(SystemSecurityPlanSerializer):

    @staticmethod
    def ssp_statement_id_from_control(control_id, part_id):
        if part_id:
            return f"{control_id}_smt.{part_id}"
        else:
            return f"{control_id}_smt"

    def create_statement_dicts(self, control_id, group_impl_smts):
        """
        Create statements for each group of implementation statements
        """
        statements = []
        for smt in group_impl_smts:
            # Get oscal control and form statement id with part if present
            cl_id = oscalize_control_id(control_id)
            smt_id = self.ssp_statement_id_from_control(cl_id, smt.pid)
            # TODO: "set-parameters" ?
            statement_dict = {"statement-id": smt_id, "uuid": str(uuid.uuid4()),
                              "by-components": [{
                                  "component-uuid": str(smt.producer_element.uuid),
                                  "uuid": str(smt.uuid),
                                  "description": smt.body}
                              ]}
            statements.append(statement_dict)
        return statements

    def as_json(self):

        # Build OSCAL SSP
        # Example: https://github.com/usnistgov/oscal-content/tree/master/examples/ssp/json/ssp-example.json
        project = self.system.projects.first()
        # TODO: again a stub as found in module_logic, this is still not entirely valid.
        profile = urlunparse((GOVREADY_URL.scheme, GOVREADY_URL.netloc,
                              "profile_path",
                              None, None, None))
        orgs = list(Organization.objects.filter(projects=project))  # TODO: orgs need uuids
        components = Element.objects.filter(statements_produced__in=self.impl_smts).filter(component_state="operational").exclude(element_type='system')
        impl_comps = [{ "component-uuid": str(component.uuid) } for component in components]
        parties = [{"uuid":str(uuid.uuid4()), "type": "organization", "name": org.name} for org in orgs]
        of = {
            "system-security-plan": {
                "uuid": str(self.system.root_element.uuid),
                "metadata": {
                    "title": "{} System Security Plan".format(self.system.root_element.name),
                    "last-modified": self.system.root_element.updated.replace(microsecond=0).isoformat(),
                    "version": project.version,
                    "oscal-version": self.system.root_element.oscal_version,
                    "roles": [],
                    "parties": parties,
                },
                "import-profile": {
                    "href": profile
                },
                "system-characteristics": {},
                "system-implementation": {
                    "remarks": "",
                    "users": [],
                    "components": [],
                    "inventory-items": [        {
          "uuid": str(uuid.uuid4()),
          "description": "An inventory item",
          "implemented-components": impl_comps
        }]# TODO: inventory-items props, description, responsible-parties
                },
                "control-implementation": {
                    "description": "",
                    "implemented-requirements": [], # implemented-requirements
                }
            }
        }
        # Create a list of dicts that are the implementation requirements
        # Each element is for each control sid group
        # Each group has that controls statements
        for control_id, group in groupby(natsorted(self.impl_smts, key=lambda ismt: ismt.sid),
                                         lambda ismt: ismt.sid):
                imp_req_dict = {
                    "uuid": str(uuid.uuid4()),
                    "control-id": "{}".format(control_id),
                    "statements":  self.create_statement_dicts(control_id, group)
                }  #statements

                of["system-security-plan"]["control-implementation"]["implemented-requirements"].append(imp_req_dict)

        # System implementation
        users = project.get_all_participants()# TODO:Need proper user title based on is_member, is_admin, editor_of
        # TODO: party-uuids users don't have uuids not sure what to do other than make a random one
        user_party_uuid = str(uuid.uuid4())

        of["system-security-plan"]["metadata"]['roles'] =  [{"id": auths.get('role', "member").split(';')[0], "title": auths.get('role', "member").split(';')[0].capitalize()  } for user, auths in users]
        of["system-security-plan"]["system-implementation"]['users'] = [{"uuid":user_party_uuid, "title":user.username, "role-ids": [auths.get('role', "member").split(';')[0]]} for user, auths in users]
        of["system-security-plan"]["system-implementation"]['components'] = [
            {
                "uuid":str(comp_ele.uuid),
                "title":comp_ele.name,
                "description":comp_ele.description,
                "status": {
                    "state": comp_ele.component_state
                },
                "type":comp_ele.component_type,
                "responsible-roles": [
                    {
                        "role-id": "asset-owner",
                        "party-uuids": [user_party_uuid]
                    }
                ],
                "props": [
                    {
                        "name": "tag",
                        "ns": "https://govready.com/ns/oscal",
                        "value": tag.label
                    } for tag in comp_ele.tags.all()
                ]
            } for comp_ele in components
        ]
        # TODO: responsible-roles
        
        # System characteristics
        # TODO: status remarks, authorization-boundary
        security_body = project.system.get_security_impact_level
        confidentiality = security_body.get("security_objective_confidentiality", "UNKNOWN")
        integrity = security_body.get("security_objective_integrity", "UNKNOWN")
        availability = security_body.get("security_objective_availability", "UNKNOWN")
        information_types = [
            {
                "uuid": str(uuid.uuid4()),
                "title": "UNKNOWN information type title",
                # "categorizations": [], # TODO https://doi.org/10.6028/NIST.SP.800-60v2r1
                "description": "information type description",
                "confidentiality-impact":  {
              "base": confidentiality
            },
                "integrity-impact":  {
              "base": integrity
            },
                "availability-impact":  {
              "base": availability
            }
            }
        ]
        
        of["system-security-plan"]["system-characteristics"] = {
            "system-name": self.system.root_element.name,
            "description": self.system.root_element.description,
            "system-ids": [
                {
                    "id": str(self.system.root_element.uuid), # TODO: identifier-type
                    "identifier-type": "https://ietf.org/rfc/rfc4122"
                }
            ],
            "security-sensitivity-level": self.system.get_security_sensitivity_level if self.system.get_security_sensitivity_level else "UNKNOWN",
            "system-information": {
                "information-types": information_types
            },
            "security-impact-level": {
                "security-objective-confidentiality": confidentiality,
                "security-objective-integrity": integrity,
                "security-objective-availability": availability
            }, 
            "status": {
                "state": self.system.root_element.component_state,
                "remarks": ""
            }, 
            "authorization-boundary": {
                "description": "The description of the authorization boundary would go here."
            }
        }
        try:
            # Create a temporary directory and dump the json_object in there.
            tempdir = tempfile.mkdtemp()
            path = os.path.join(tempdir, "ssp_object.json")
            # Use trestle's ComponentDefinition method oscal_read to read the path to json in the temporary folder
            path_ssp_definition = pathlib.Path(path)

            with open(path, 'w+') as cred:
                json.dump(of, cred)
            # Read in temporary file and shape into trestle pydantic SSP definition.
            trestle_oscal_json = trestlessp.SystemSecurityPlan.oscal_read(path_ssp_definition)
            # Finally validate that this object is valid by the OSCAL System Security Plan definition
            trestlessp.SystemSecurityPlan.validate(trestle_oscal_json)
            # Cleanup
            shutil.rmtree(tempdir)
        except Exception as e:
            logger.error(f"Invalid System Security Plan JSON: {e}")
            shutil.rmtree(tempdir)
            return HttpResponse(e)

        oscal_string = json.dumps(of, sort_keys=False, indent=2)
        return oscal_string

class ComponentSerializer(object):

    def __init__(self, element, impl_smts):
        self.element = element
        self.impl_smts = impl_smts

class OSCALComponentSerializer(ComponentSerializer):

    @staticmethod
    def statement_id_from_control(control_id, part_id):
        # Checking for a case where the control was provided like ac-2.3 which already has its part included.
        if part_id:
            if part_id not in control_id:
                return f"{control_id}.{part_id}"

        return f"{control_id}"

    def generate_source(self, src_str):
        """Return a valid catalog source given string"""
        DEFAULT_SOURCE = "NIST_SP-800-53_rev5"
        if not src_str:
            return DEFAULT_SOURCE
        # TODO: Handle other cases
        source = src_str
        return source

    def as_json(self):
        # Build OSCAL
        # Example: https://github.com/usnistgov/OSCAL/blob/master/src/content/ssp-example/json/example-component.json
        comp_uuid = str(self.element.uuid)
        control_implementations = []
        props = []
        orgs = list(Organization.objects.all())  # TODO: orgs need uuids, not sure which orgs to use for a component
        parties = [{"uuid": str(uuid.uuid4()), "type": "organization", "name": org.name} for org in orgs]
        responsible_roles =  [{
           "role-id": "supplier",# TODO: Not sure what this refers to
            "party-uuids": [ str(party.get("uuid")) for party in parties ]

        }]
        of = {
            "component-definition": {
                "uuid": str(uuid4()),
                "metadata": {
                    "title": "{}".format(self.element.name),
                    "last-modified": self.element.updated.replace(microsecond=0).isoformat(),
                    "version": self.element.updated.replace(microsecond=0).isoformat(),
                    "oscal-version": self.element.oscal_version,
                    "parties": parties
                },
                "components": [
                   {
                        "uuid": comp_uuid,
                        "type": self.element.component_type.lower() if self.element.component_type is not None else "software",
                        "title": self.element.full_name or self.element.name,
                        "description": self.element.description,
                        "responsible-roles": responsible_roles, # TODO: gathering party-uuids, just filling for now
                        "props": props,
                        "control-implementations": control_implementations
                    }
                ]
            },
        }

        # Add component's tags if they exist
        if self.element.tags.exists():
            props.extend([{"name": "tag", "ns": "https://govready.com/ns/oscal", "value": tag.label} for tag in self.element.tags.all()])

        # Remove 'metadata.props' key if no metadata.props exist
        if len(props) == 0:
            of['component-definition']['metadata'].pop('props', None)

        # create requirements and organize by source (sid_class)

        by_class = defaultdict(list)

        # work:
        # group stmts by control-id
        # emit an requirement for the control-id
        # iterate over each group
        # emit a statement for each member of the group
        # notes:
        # - OSCAL implemented_requirements and control_implementations need UUIDs
        #   which we don't have in the db, so we construct them.

        for control_id, group in groupby(natsorted(self.impl_smts, key=lambda ismt: ismt.sid),
                                         lambda ismt: ismt.sid):

            for smt in group:
                statement_id = self.statement_id_from_control(control_id, smt.pid)
                statement_req = {
                    "uuid": str(smt.uuid),
                    "description": smt.body,
                    "control-id": statement_id,
                }
                # key-value by sid a.k.a control id for each requirement
                if statement_req not in by_class[smt.sid]:
                    by_class[smt.sid].append(statement_req)

        for sid_class, requirements in by_class.items():
            control_implementation = {
                "uuid":str(uuid4()),# TODO: Not sure if this should implemented or just generated here.
                "source": self.generate_source(smt.source if smt.source else None),
                "description": f"This is a partial implementation of the {sid_class} catalog, focusing on the control enhancement {requirements[0].get('control-id')}.",
                "implemented-requirements": [req for req in requirements]
            }
            control_implementations.append(control_implementation)
        # Remove 'control-implementations' key if no implementations exist
        if len(control_implementations) == 0:
            of['component-definition']['components'][0].pop('control-implementations', None)

        oscal_string = json.dumps(of, sort_keys=False, indent=2)
        return oscal_string

class OpenControlComponentSerializer(ComponentSerializer):

    def as_yaml(self):
        ocf = {
                "name": self.element.name,
                "schema_version": "3.0.0",
                "documentation_complete": False,
                "satisfies": []
               }

        satisfies_smts = ocf["satisfies"]
        for smt in self.impl_smts:
            my_dict = {
                "control_key": smt.sid.upper(),
                "control_name": smt.catalog_control_as_dict['title'],
                "standard_key": smt.sid_class,
                "covered_by": [],
                "security_control_type": "Hybrid | Inherited | ...",
                "narrative": [
                    {"text": smt.body}
                ],
                "remarks": [
                    {"text": smt.remarks}
                ]
            }
            satisfies_smts.append(my_dict)
        opencontrol_string = rtyaml.dump(ocf)
        return opencontrol_string

class ComponentImporter(object):

    def import_components_as_json(self, import_name, json_object, request=None, existing_import_record=False, stopinvalid=True):
        """Imports Components from a JSON object

        @type import_name: str
        @param import_name: Name of import file (if it exists)
        @type json_object: dict
        @param json_object: Element attributes from JSON object
        @type existing_import_record: boolean
        @param existing_import_record: Continue to append imports to an existing import record
        @rtype: ImportRecord if success, bool (false) if failure
        @returns: ImportRecord linked to the created components (if success) or False if failure
        """
        issues = []
        try:
            # Create a temporary directory and dump the json_object in there.
            tempdir = tempfile.mkdtemp()
            path = os.path.join(tempdir, "object.json")
            # Use trestle's ComponentDefinition method oscal_read to read the path to json in the temporary folder
            path_component_definition = pathlib.Path(path)
            oscal_json = json.loads(json_object)
            with open(path, 'w+') as cred:
                json.dump(oscal_json, cred)
            # Read in temporary file and shape into trestle pydantic Component definition.
            trestle_oscal_json = trestlecomponent.ComponentDefinition.oscal_read(path_component_definition)
            # Finally validate that this object is valid by the component definition
            trestlecomponent.ComponentDefinition.validate(trestle_oscal_json)
            # Cleanup
            shutil.rmtree(tempdir)
        except Exception as e:
            logger.error(e)
            logger.warning(
                event="error_importing_component",
                object={"object": "component", "name": import_name, "error": {e.__context__}}
                )
            issues.append({"object": "component", "name": import_name, "error": {e.__context__}})
            shutil.rmtree(tempdir)
            # Check a component uploaded to form
            if request and request.POST.get("json_content") is not None:
                if stopinvalid:
                    messages.add_message(request, messages.ERROR, f"IMPORT HALTED. Invalid Component JSON: {e.__context__}")
                    return HttpResponse(e)
                else:
                    messages.add_message(request, messages.INFO, f"IMPORT CONTINUED WITH POSSIBLE ERROR. Invalid Component JSON: {e.__context__}")
            else:
                if stopinvalid:
                    print("\nNOTICE - ISSUES DURING COMPONENT IMPORT\n")
                    [print(issue) for issue in issues]
                    print("\nPROGRAM HALTED\n")
                    import sys
                    sys.exit()


        # If importing from importcomponents script print issues
        if len(issues) > 0:
            print("\nNOTICE - ISSUES DURING COMPONENT IMPORT\n")
            [print(issue) for issue in issues]

        # Returns list of created components
        if request is not None:
            user_owner = request.user
        else:
            user_owner = User.objects.filter(is_superuser=True)[0]
        created_components = self.create_components(oscal_json, user_owner)
        new_import_record = self.create_import_record(import_name, created_components, existing_import_record=existing_import_record)
        return new_import_record

    def create_import_record(self, import_name, components, existing_import_record=False):
        """Associates components and statements to an import record

        @type import_name: str
        @param import_name: Name of import file (if it exists)
        @type components: list
        @param components: List of components
        @type existing_import_record: booleen
        @param existing_import_record: Continue to append imports to an existing import record
        @rtype: ImportRecord
        @returns: New ImportRecord object with components and statements associated
        """
        import_record = ImportRecord.objects.filter(name=import_name).last()
        if import_record is None or not existing_import_record:
            import_record = ImportRecord.objects.create(name=import_name)
        for component in components:
            statements = Statement.objects.filter(producer_element=component)
            for statement in statements:
                statement.import_record.add(import_record)
                #statement.save()
            component.import_record = import_record
            component.save()

        return import_record

    def create_components(self, oscal_json, user_owner=None):
        """Creates Elements (Components) from valid OSCAL JSON"""
        components_created = []
        components = oscal_json['component-definition']['components']
        for component in components:
            new_component = self.create_component(component, user_owner)
            if new_component is not None:
                components_created.append(new_component)

        return components_created

    def create_component(self, component_json, user_owner=None, private=False):
        """Creates a component from a JSON dict

        @type component_json: dict
        @param component_json: Component attributes from JSON object
        @param user_owner: Django user
        @rtype: Element
        @returns: Element object if created, None otherwise
        """

        component_name = component_json['title']
        
        while Element.objects.filter(name=component_name).count() > 0:
            component_name = increment_element_name(component_name)

        new_component = Element.objects.create(
            name=component_name,
            description=component_json['description'] if 'description' in component_json else 'Description missing',
            # Components uploaded to the Component Library are all system_element types
            element_type="system_element",
            uuid=component_json['uuid'] if 'uuid' in component_json else uuid.uuid4(),
            component_type=component_json['type'] if 'type' in component_json else "software",
            private=private
        )

        logger.info(f"Component {new_component.name} created with UUID {new_component.uuid}.")
        if user_owner:
            new_component.assign_owner_permissions(user_owner)
            logger.info(
                event="new_element with user as owner",
                object={"object": "element", "id": new_component.id, "name":new_component.name},
                user={"id": user_owner.id, "username": user_owner.username}
            )

        component_props = component_json.get('props', None)
        if component_props:
            desired_tags = set([prop['value'] for prop in component_props if prop['name'] == 'tag' and 'ns' in prop and prop['ns'] == "https://govready.com/ns/oscal"])
            existing_tags = Tag.objects.filter(label__in=desired_tags).values('id', 'label')
            tags_to_create = desired_tags.difference(set([tag['label'] for tag in existing_tags]))
            new_tags = Tag.objects.bulk_create([Tag(label=tag) for tag in tags_to_create])
            all_tag_ids = [tag.id for tag in new_tags] + [tag['id'] for tag in existing_tags]
            new_component.add_tags(all_tag_ids)
            new_component.save()
        created_statements = []
        control_implementation_statements = component_json.get('control-implementations', None)
        # If there data exists the OSCAL component's control-implementations key
        if control_implementation_statements:
            for control_element in control_implementation_statements:
                catalog = oscalize_catalog_key(control_element.get('source', None))
                created_statements = self.create_control_implementation_statements(catalog, control_element, new_component)
        # If there are no valid statements in the json object
        if created_statements == []:
            logger.info(f"The Component {new_component.name} will be deleted as there were no valid statements provided.")
            new_component.delete()
            new_component = None

        return new_component

    def create_control_implementation_statements(self, catalog_key, control_element, parent_component):
        """Creates a Statement from a JSON dict implemented-requirements

        @type catalog_key: str
        @param catalog_key: Catalog of the control statements
        @type control_element: dict
        @param control_element: Implemented controls
        @type parent_component: str
        @param parent_component: UUID of parent component
        @rtype: dict
        @returns: New statement objects created
        """

        new_statements = []
        implemented_reqs = control_element['implemented-requirements'] if 'implemented-requirements' in control_element else []
        for implemented_control in implemented_reqs:
            control_id = implemented_control['control-id'] if 'control-id' in implemented_control else 'missing'
            new_statement = Statement(
                sid=control_id,
                sid_class=catalog_key,
                pid=get_control_statement_part(control_id),
                source=catalog_key,
                uuid=control_element['uuid'] if 'uuid' in control_element else uuid.uuid4(),
                body=implemented_control['description'] if 'description' in implemented_control else '',
                statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
                remarks=implemented_control['remarks'] if 'remarks' in implemented_control else '',
                status=implemented_control['status'] if 'status' in implemented_control else None,
                producer_element=parent_component,
            )
            logger.info(f"New statement with UUID {new_statement.uuid} being created.")
            new_statements.append(new_statement)
        statements_created = Statement.objects.bulk_create(new_statements)
        return statements_created

def add_selected_components(system, import_record):
        """Add a component from the library or a compliance app to the project and its statements using the import record"""

        # Get components from import record
        imported_components = Element.objects.filter(import_record=import_record)
        for imported_component in imported_components:
            # Loop through all element's prototype statements and add to control implementation statements.
            # System's selected controls will filter what controls and control statements to display.
            for smt in Statement.objects.filter(producer_element_id=imported_component.id,
                                                statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name):
                smt.create_system_control_smt_from_component_prototype_smt(system.root_element.id)
        return imported_components

@login_required
def system_element(request, system_id, element_id):
    """Display System's selected element detail view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        # Retrieve element
        element = Element.objects.get(id=element_id)
        # Retrieve impl_smts produced by element and consumed by system
        # Get the impl_smts contributed by this component to system
        # if this is a proposal then we wont have any impl_smts, so we need to check if there is a proposal
        # and if impl_smts is empty
        
        impl_smts = element.statements_produced.filter(consumer_element=system.root_element)
        
        if(impl_smts.exists()):
            # Retrieve used catalog_key
            catalog_key = impl_smts[0].sid_class     
            component_request = None
            hasSentRequest = False
            # Retrieve related Proposal and request
            try:
                proposals = Proposal.objects.filter(requested_element=element.id, req__system=system, req__status="Approve")
                proposal = proposals.first()
                if proposal != None:
                    component_request = proposal.req
                    hasSentRequest = True
            except Proposal.DoesNotExist:
                proposal = None

            # Retrieve control ids
            catalog_controls = Catalog.GetInstance(catalog_key=catalog_key).get_controls_all()

            # Build OSCAL and OpenControl
            oscal_string = OSCALComponentSerializer(element, impl_smts).as_json()
            opencontrol_string = OpenControlComponentSerializer(element, impl_smts).as_yaml()
            states = [choice_tup[1] for choice_tup in ComponentStateEnum.choices()]
            types = [choice_tup[1] for choice_tup in ComponentTypeEnum.choices()]
            # Return the system's element information
            
            context = {
                "states": states,
                "types": types,
                "system": system,
                "project": project,
                "element": element,
                "impl_smts": impl_smts,
                "catalog_controls": catalog_controls,
                "catalog_key": catalog_key,
                "oscal": oscal_string,
                "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
                "opencontrol": opencontrol_string,
                "display_urls": project_context(project),
                "proposal": proposal,
                "hasSentRequest": hasSentRequest,
                "component_request": component_request,
            }
            return render(request, "systems/element_detail_tabs.html", context)
        else:
            try:
                proposal = system.proposals.get(requested_element__id=element_id)
            except Proposal.DoesNotExist: 
                proposal = None
                return HttpResponseRedirect("/controls/{}/components/selected".format(system_id))

            # Get all statements that are not component_approval_criteria
            impl_smts = element.statements_produced.filter(~Q(statement_type='COMPONENT_APPROVAL_CRITERIA'))
            # Retrieve used catalog_key
            catalog_key = impl_smts[0].sid_class
            
            # Retrieve control ids
            catalog_controls = Catalog.GetInstance(catalog_key=catalog_key).get_controls_all()
            # Build OSCAL and OpenControl
            oscal_string = OSCALComponentSerializer(element, impl_smts).as_json()
            opencontrol_string = OpenControlComponentSerializer(element, impl_smts).as_yaml()
            states = [choice_tup[1] for choice_tup in ComponentStateEnum.choices()]
            types = [choice_tup[1] for choice_tup in ComponentTypeEnum.choices()]
            # Return the system's element information
            
            requests = Request.objects.filter(system=system, requested_element=element)
            component_request = None
            hasSentRequest = False
            
            if proposal.req != None:
                hasSentRequest = True
                component_request = proposal.req

            context = {
                "states": states,
                "types": types,
                "system": system,
                "project": project,
                "element": element,
                "proposal": proposal,
                "component_request": component_request,
                "hasSentRequest": hasSentRequest,
                "impl_smts": impl_smts,
                "catalog_controls": catalog_controls,
                "catalog_key": catalog_key,
                "oscal": oscal_string,
                "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
                "opencontrol": opencontrol_string,
                "display_urls": project_context(project)
            }
            return render(request, "systems/element_detail_tabs.html", context)

@login_required
def system_element_control(request, system_id, element_id, catalog_key, control_id):
    """Display System's selected element detail view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        # Retrieve element
        element = Element.objects.get(id=element_id)

        # Retrieve impl_smts produced by element and consumed by system
        # Get the impl_smts contributed by this component to system
        impl_smts = element.statements_produced.filter(consumer_element=system.root_element)
        # Get the cont
        impl_smt_ctl = next((ctl for ctl in impl_smts if ctl.sid == control_id and ctl.sid_class == catalog_key), None)

        # Retrieve control ids
        # TODO: Only need to individual control
        catalog_controls = Catalog.GetInstance(catalog_key=catalog_key).get_controls_all()
        # Retrieve control
        control = next((ctl for ctl in catalog_controls if ctl['id'] == oscalize_control_id(control_id)), None)

        # Build OSCAL and OpenControl
        oscal_string = OSCALComponentSerializer(element, impl_smts).as_json()
        opencontrol_string = OpenControlComponentSerializer(element, impl_smts).as_yaml()
        states = [choice_tup[1] for choice_tup in ComponentStateEnum.choices()]
        types = [choice_tup[1] for choice_tup in ComponentTypeEnum.choices()]
        # Return the system's element information
        context = {
            "states": states,
            "types": types,
            "system": system,
            "project": project,
            "element": element,
            "impl_smts": impl_smts,
            "impl_smt_ctl": impl_smt_ctl,
            "catalog_controls": catalog_controls,
            "catalog_key": catalog_key,
            "control": control,
            "oscal": oscal_string,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "opencontrol": opencontrol_string,
        }
        return render(request, "systems/element_detail_control.html", context)

def edit_component_state(request, element_id):
    """
    Edit system component state
    """

    # Retrieve element
    # TODO: Make atomic transaction
    element = Element.objects.get(id=element_id)
    element.component_state = request.POST['state_change']
    element.save()
    logger.info(event=f"change_system update_component_state {element} {element.component_state}",
                object={"object": "element", "id": element.id},
                user={"id": request.user.id, "username": request.user.username})
    # Batch update status of control implementation statements provided by the element to the system
    state_status = {"operational": "Implemented", "under-development": "Partially Implemented", "planned": "Planned"}
    control_status = state_status.get(request.POST['state_change']) or "Not Implemented"
    return redirect(reverse('component_library_component', args=[element_id]))

def edit_component_type(request, element_id):
    """
    Edit component type
    """

    # Retrieve element
    element = Element.objects.get(id=element_id)
    element.component_type = request.POST['type_change']
    element.save()
    return redirect(reverse('component_library_component', args=[element_id]))

@login_required
def system_element_remove(request, system_id, element_id):
    """Remove an element from a system and delete the controls it produces"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('change_system', system):

        # Retrieve element
        element = Element.objects.get(id=element_id)
        try:
            req = Request.objects.get(system=system, requested_element=element, status="Approve")
            req.status = "Closed"
            req.save()
        except Request.DoesNotExist:
            pass
        
        # Delete the control implementation statements associated with this component
        result = element.statements_produced.filter(consumer_element=system.root_element).delete()

        # Log result
        logger.info(
                event="change_system remove_component",
                object={"object": "component", "id": element.id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        messages.add_message(request, messages.INFO, f"Removed component '{element.name}' from system.")

    else:
        # User does not have permission
        # Log result
        logger.info(
                event="change_system remove_component permission_denied",
                object={"object": "component", "id": element_id},
                user={"id": request.user.id, "username": request.user.username}
                )

        # Create message for user
        messages.add_message(request, messages.INFO, f"You do not have permission to edit the system.")

    response = redirect(reverse('components_selected', args=[system_id]))
    return response

@login_required
def new_element(request):
    """Form to create new system element (aka component)"""

    if request.method == 'POST':
        form = ElementForm(request.POST)
        if form.is_valid():
            form.save()
            
            element = form.instance
            element.assign_owner_permissions(request.user)

            Statement.objects.create(
                sid = None,
                sid_class = None,
                body = "",
                statement_type = StatementTypeEnum.COMPONENT_APPROVAL_CRITERIA.name,
                producer_element = element
            )
            
            logger.info(
                event="new_element with user as owner",
                object={"object": "element", "id": element.id, "name":element.name},
                user={"id": request.user.id, "username": request.user.username}
            )
            
            return redirect('component_library_component', element_id=element.id)
    else:
        form = ElementForm()

    return render(request, 'components/element_form.html', {
        'form': form,
    })

@login_required
def edit_element_access_management(request, element_id):
    """Form to edit system element access management"""

    # The original element(component) 
    element = get_object_or_404(Element, id=element_id)
    # statement related to element
    if element.statements_produced.filter(statement_type=StatementTypeEnum.COMPONENT_APPROVAL_CRITERIA.name).exists():
        statement = element.statements_produced.filter(statement_type=StatementTypeEnum.COMPONENT_APPROVAL_CRITERIA.name).first()
    else:
        # create new statement and assign statement to element
        statement = Statement.objects.create(
            sid = None,
            sid_class = None,
            body = "",
            statement_type = StatementTypeEnum.COMPONENT_APPROVAL_CRITERIA.name,
            producer_element = element
        )
    if request.method == 'POST':
        form = ElementEditAccessManagementForm(request.POST or None, instance=element)
        statementForm = StatementEditForm(request.POST, instance=statement)
        if form.is_valid() and statementForm.is_valid():
            logger.info(
                event="edit_element_access_management",
                object={"object": "element", "id": form.instance.id, "name": form.instance.name},
                user={"id": request.user.id, "username": request.user.username}
            )

            form.save()
            statementForm.save()
            return JsonResponse({"status": "ok"})
        else:
            errors = form.errors.get_json_data(escape_html=False)
            msg_list = [f"{e.title()} - {errors[e][0]['message']}" for e in errors.keys()]
            return JsonResponse({"status": "err", "message": "Please fix the following problems:<br>"+"<br>".join(msg_list)})


@login_required
def component_library_component(request, element_id):
    """Display library component's element detail view"""
    
    # Retrieve element
    element = Element.objects.get(id=element_id)
    govSystem = System.objects.filter(root_element__name="GovReady-Q Sample System")
   # Check permissions
    if element.private == True and 'view_element' not in get_user_perms(request.user, element):
        logger.warning(
            event="view_element_private permission_denied",
            comment=f"User {request.user.username} does not have permission to view this element {element.name}",
            object={"object": "element", "id": element.id},
            user={"id": request.user.id, "username": request.user.username}
        )
        raise Http404
    hasPermissionToEdit = 'change_element' in get_user_perms(request.user, element)
    smt_query = request.GET.get('search')
    usersWithPermission = get_users_with_perms(element, attach_perms=True)
    listUsers = []
    
    for user in usersWithPermission:
        listUsers.append(user.username)

    listOfContacts = []
    poc_users = element.appointments.filter(role__role_id='poc', party__party_type='POC')


    for poc in poc_users:
        user = {
            "uuid":poc.party.uuid,
            "party_type":poc.party.party_type,
            "name":poc.party.name,
            "short_name":poc.party.short_name,
            "email":poc.party.email,
            "phone_number":poc.party.phone_number,
            "role_title":poc.role.title,
            "role_name":poc.role.short_name,
        }
        listOfContacts.append(user)

    get_all_parties = element.appointments.all()
    total_number_of_requests = element.requests.filter(status__in=["Open", "Pending", "In Progress", "Reject", "Approve"]).count()

    contacts = []
    for poc in get_all_parties:
        contacts.append(poc.party)
    
    criteria_results = element.statements_produced.filter(statement_type=StatementTypeEnum.COMPONENT_APPROVAL_CRITERIA.name)
    if len(criteria_results) > 0:
        criteria_text = criteria_results.first().body
    else:
        criteria_text = ""
    is_owner = element.is_owner(request.user)
    
    # Retrieve systems consuming element
    consuming_systems = element.consuming_systems()
    states = [choice_tup[1] for choice_tup in ComponentStateEnum.choices()]
    types = [choice_tup[1] for choice_tup in ComponentTypeEnum.choices()]

    if smt_query:
        impl_smts = element.statements_produced.filter(sid__icontains=smt_query, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name)
    else:
        # Get the impl_smts contributed by this component to system
        impl_smts = element.statements_produced.filter(statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name)

    if len(impl_smts) < 1:
        # New component, no control statements assigned yet
        catalog_key = "catalog_key_missing"
        catalog_controls = None
        oscal_string = OSCALComponentSerializer(element, impl_smts).as_json()
        opencontrol_string = None
        context = {
            "element": element,
            "states": states,
            "impl_smts": impl_smts,
            "oscal": oscal_string,
            "is_admin": request.user.is_superuser,
            "list_of_permissible_users": listUsers,
            "is_owner": is_owner,
            "can_edit": hasPermissionToEdit,
            "users_with_permissions": usersWithPermission,
            "criteria": criteria_text,
            "listOfContacts": listOfContacts,
            "contacts": serializers.serialize('json', contacts),
            "totalRequests": total_number_of_requests,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "form_source": "component_library"
        }
        return render(request, "components/element_detail_tabs.html", context)
    elif len(impl_smts) > 0:
        # TODO: We may have multiple catalogs in this case in the future
        # Retrieve used catalog_key
        catalog_key = impl_smts[0].sid_class
        # Retrieve control ids
        catalog_controls = Catalog.GetInstance(catalog_key=catalog_key).get_controls_all()
        # Build OSCAL and OpenControl
        oscal_string = OSCALComponentSerializer(element, impl_smts).as_json()
        opencontrol_string = OpenControlComponentSerializer(element, impl_smts).as_yaml()

    # Use natsort here to handle the sid that has letters and numbers
    # (e.g. to put AC-14 after AC-2 whereas before it was putting AC-14 before AC-2)
    # using the natsort package from pypi: https://pypi.org/project/natsort/
    impl_smts = natsorted(impl_smts, key=lambda x: x.sid)

    # Pagination
    obj_paginator = Paginator(impl_smts, 15)
    page_number = request.GET.get('page')

    try:
        page_obj = obj_paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = obj_paginator.page(1)
    except EmptyPage:
        page_obj = obj_paginator.page(obj_paginator.num_pages)

    # Return the system's element information
    context = {
        "page_obj": page_obj,
        "element": element,
        "states": states,
        "types": types,
        "consuming_systems": consuming_systems,
        "impl_smts": impl_smts,
        "catalog_controls": catalog_controls,
        "catalog_key": catalog_key,
        "oscal": oscal_string,
        "is_admin": request.user.is_superuser,
        "list_of_permissible_users": listUsers,
        "is_owner": is_owner,
        "can_edit": hasPermissionToEdit,
        "users_with_permissions": usersWithPermission,
        "criteria": criteria_text,
        "contacts": serializers.serialize('json', contacts),
        "totalRequests": total_number_of_requests,
        "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
        "enable_experimental_oscal": SystemSettings.enable_experimental_oscal,
        "opencontrol": opencontrol_string,
        # "project_form": AddProjectForm(request.user, initial={'portfolio': request.user.portfolio_list().first().id}),
        "tags": json.dumps(SimpleTagSerializer(element.tags, many=True).data),
        "form_source": "component_library"
    }
    return render(request, "components/element_detail_tabs.html", context)

@login_required
def api_controls_select(request):
    """Return list of controls in json for select2 options from all control catalogs"""

    cl_id = request.GET.get('q', None)
    oscal_ctl_id = oscalize_control_id(cl_id)
    catalogs_containing_cl_id = CatalogData.objects.filter(Q(catalog_json__catalog__groups__contains=[{"controls": [{"id": oscal_ctl_id}]}]) |
        Q(catalog_json__catalog__groups__contains=[{"controls": [{"controls": [{"id": oscal_ctl_id}]}] }] ))
    cxs = []
    for catalog in catalogs_containing_cl_id:
        catalog_key_display = catalog.catalog_key.replace("_", " ")
        # TODO: get control title effectively from CatalogData
        # title = "catalog.ctl_title"
        cxs.append({"id": oscal_ctl_id, 'catalog_key_display': catalog_key_display, 'display_text': f"{oscal_ctl_id} - {catalog_key_display} - {cl_id}"})
    status = "success"
    message = "Sending list."
    return JsonResponse( {"status": status, "message": message, "data": {"controls": cxs} })

@login_required
def component_library_component_copy(request, element_id):
    """Copy a component"""

    # Retrieve element
    element = Element.objects.get(id=element_id)
    count = Element.objects.filter(uuid=element.uuid).count()

    if count > 0:
        e_copy = element.copy(name=element.name + " copy ("+str(count+1)+')')
    else:
        e_copy = element.copy()
    e_copy.assign_owner_permissions(request.user)
    # Create message to display to user
    messages.add_message(request, messages.INFO,
                         'Component "{}" copied to "{}".'.format(element.name, e_copy.name))

    # # Redirect to the new page for the component
    return HttpResponseRedirect("/controls/components/{}".format(e_copy.id))

@login_required
def import_component(request):
    """Import a Component in JSON"""

    import_name = request.POST.get('import_name', '')
    oscal_component_json = request.POST.get('json_content', '')
    result = ComponentImporter().import_components_as_json(import_name, oscal_component_json, request)
    return component_library(request)

def raise_404_if_not_permitted_to_statement(request, statement, system_permission='view_system'):
    """Raises a 404 if the user doesn't have the statement system permission"""
    while True:
        if request.user.is_superuser or request.user.is_staff:
            # Allow access if the user has superuser or staff member permissions respectively
            break
        else:
            # Allow access if the user has the statement system permission
            try:
                system = System.objects.get(root_element=statement.consumer_element)
                if request.user.has_perm(system_permission, system):
                    break
            except System.DoesNotExist:
                logger.info(
                event="update_smt_permission permission_denied",
                object={"object": "statement", "id": statement.id},
                user={"id": request.user.id, "username": request.user.username}
                )
                raise Http404

@login_required
def statement_history(request, smt_id=None):
    """Returns the history for the specified statement"""

    # Get statement if exists else 404
    smt = get_object_or_404(Statement, id=smt_id)


    # Check permission block
    permission = False
    if request.user.is_superuser or request.user.is_staff:
        # Grant permission to superusers and staff users
        permission = True
    elif System.objects.filter(root_element=smt.consumer_element).exists():
        # Grant permission to user with edit access on system
        system = System.objects.get(root_element=smt.consumer_element)
        permission = True if request.user.has_perm('view_system', system) else False
    # 404 if user does not have permission
    if not permission:
        raise Http404
    # Check permission block end

    full_smt_history = None
    try:
        full_smt_history = smt.history.all()
    except Statement.DoesNotExist:
        messages.add_message(request, messages.ERROR, f'The statement id is not valid. Is this still a statement in GovReady?')


    # Check permission
    raise_404_if_not_permitted_to_statement(request, smt)

    return render(request, "controls/statement_history.html", {"statement": smt.history.all()})

@login_required
def restore_to_history(request, smt_id, history_id):
    """Restores the specified statement version"""

    # Get statement if exists else 404
    smt = get_object_or_404(Statement, id=smt_id)

    print(1, "==== smt", smt)

    # Check permission
    raise_404_if_not_permitted_to_statement(request, smt, 'change_system')

    full_smt_history = None
    for query_key in request.POST:
        if "restore" in query_key:
            change_reason = request.POST.get(query_key, "")
        else:
            change_reason = None

    try:
        # Save historical statement as a new instance
        historical_smt = smt.history.get(history_id=history_id)
        historical_smt.instance.save()

        # Update the reason for the new statement record
        recent_smt = smt.history.first()
        # update_change_reason(recent_smt.instance, change_reason)

        logger.info( f"Change reason: {change_reason}")

        logger.info(
            f"Restoring the current statement with an id of {smt_id} to version with a history id of {history_id}")
        messages.add_message(request, messages.INFO,
                             f'Successfully restored the statement to version history {history_id}')

        # Diff between most recent and the historical record
        full_smt_history = smt.history.all()
        recent_record = full_smt_history.filter(history_id=recent_smt.history_id).first()
        historical_record = full_smt_history.filter(history_id=historical_smt.history_id).first()

        delta = historical_record.diff_against(recent_record)
        for change in delta.changes:
            logger.info("{} changed from {} to {}".format(change.field, change.old, change.new))
    except ObjectDoesNotExist as ex:
        messages.add_message(request, messages.ERROR, f'{ex} Is this still a statement record in GovReady?')

    context = {
        "history_id": history_id,
        "smt_id": smt_id,
        "statement": full_smt_history}

    return render(request, "controls/statement_history.html", context)

@login_required
def system_element_download_oscal_json(request, system_id, element_id):

    if system_id is not None and system_id != '':
        # Retrieve identified System
        system = System.objects.get(id=system_id)
        # Retrieve related selected controls if user has permission on system
        if request.user.has_perm('view_system', system):
            # Retrieve primary system Project
            # Temporarily assume only one project and get first project
            project = system.projects.all()[0]

            # Retrieve element
            element = Element.objects.get(id=element_id)

            # Retrieve impl_smts produced by element and consumed by system
            # Get the impl_smts contributed by this component to system
            impl_smts = element.statements_produced.filter(consumer_element=system.root_element)
    else:
        # Comes from Component Library, no system

        # Retrieve element
        element = Element.objects.get(id=element_id)
        # Get the impl_smts contributed by this component to system
        impl_smts = Statement.objects.filter(producer_element=element)
    print(f"component library statements {len(impl_smts)}")
    response = HttpResponse(content_type="application/json")
    filename = str(PurePath(slugify(element.name)).with_suffix('.json'))
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    body = OSCALComponentSerializer(element, impl_smts).as_json()
    response.write(body)

    return response

@login_required
def OSCAL_ssp_export(*args, **kwargs):
    """
    Exporting a system security plan in OSCAL json version 1.0.0
    """
    system_id = kwargs.get('system_id', 1)
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    impl_smts = Statement.objects.filter(consumer_element=system.root_element, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name)
    oscal_string = OSCALSystemSecurityPlanSerializer(system, impl_smts).as_json()
    # File name construction and JSON response
    filename = "{}_OSCAL_{}.json".format(system.root_element.name.replace(" ", "_"),
                                                           datetime.now().strftime("%Y-%m-%d-%H-%M"))
    resp = HttpResponse(oscal_string, content_type="application/json")
    resp["content-disposition"] = "attachment; filename=%s" % quote(filename)
    return resp

@login_required
def controls_selected_export_xacta_xslx(request, system_id):
    """Export System's selected controls compatible with Xacta 360"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()

        # Retrieve any related Implementation Statements
        impl_smts = system.root_element.statements_consumed.all()
        impl_smts_by_sid = {}
        for smt in impl_smts:
            if smt.sid in impl_smts_by_sid:
                impl_smts_by_sid[smt.sid].append(smt)
            else:
                impl_smts_by_sid[smt.sid] = [smt]

        for control in controls:
            if control.oscal_ctl_id in impl_smts_by_sid:
                setattr(control, 'impl_smts', impl_smts_by_sid[control.oscal_ctl_id])
            else:
                setattr(control, 'impl_smts', None)

        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
        from tempfile import NamedTemporaryFile

        wb = Workbook()
        ws = wb.active
        # create alignment style
        wrap_alignment = Alignment(wrap_text=True)
        ws.title = "Controls_Implementation"

        # Add in field name row
        # Paragraph/ReqID
        c = ws.cell(row=1, column=1, value="Paragraph/ReqID")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(left=Side(border_style="thin", color="444444"), right=Side(border_style="thin", color="444444"),
                          bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Stated Requirement (Control statement/Requirement)
        c = ws.cell(row=1, column=2, value="Title")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions['B'].width = 30
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Private Implementation
        c = ws.cell(row=1, column=3, value="Private Implementation")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions['C'].width = 80
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Public Implementation
        c = ws.cell(row=1, column=4, value="Public Implementation")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions['D'].width = 80
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Notes
        c = ws.cell(row=1, column=5, value="Notes")
        ws.column_dimensions['E'].width = 60
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Status ["Implemented", "Planned"]
        c = ws.cell(row=1, column=6, value="Status")
        ws.column_dimensions['F'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Expected Completion (expected implementation)
        c = ws.cell(row=1, column=7, value="Expected Completion")
        ws.column_dimensions['G'].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Class ["Management", "Operational", "Technical",
        c = ws.cell(row=1, column=8, value="Class")
        ws.column_dimensions['H'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Priority ["p0", "P1", "P2", "P3"]
        c = ws.cell(row=1, column=9, value="Priority")
        ws.column_dimensions['I'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Responsible Entities
        c = ws.cell(row=1, column=10, value="Responsible Entities")
        ws.column_dimensions['J'].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Control Owner(s)
        c = ws.cell(row=1, column=11, value="Control Owner(s)")
        ws.column_dimensions['K'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Type ["System-Specific", "Hybrid", "Inherited", "Common", "blank"]
        c = ws.cell(row=1, column=12, value="Type")
        ws.column_dimensions['L'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Inherited From
        c = ws.cell(row=1, column=13, value="Inherited From")
        ws.column_dimensions['M'].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Provide As ["Do Not Share", "blank"]
        c = ws.cell(row=1, column=14, value="Provide As")
        ws.column_dimensions['N'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Evaluation Status ["Evaluated", "Expired", "Not Evaluated", "Unknown", "blank"]
        c = ws.cell(row=1, column=15, value="Evaluation Status")
        ws.column_dimensions['O'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # Control Origination
        c = ws.cell(row=1, column=16, value="Control Origination")
        ws.column_dimensions['P'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        # History
        c = ws.cell(row=1, column=17, value="History")
        ws.column_dimensions['Q'].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(right=Side(border_style="thin", color="444444"), bottom=Side(border_style="thin", color="444444"),
                          outline=Side(border_style="thin", color="444444"))

        for row in range(2, len(controls) + 1):
            control = controls[row - 2]

            # Paragraph/ReqID
            c = ws.cell(row=row, column=1, value=control.get_flattened_oscal_control_as_dict()['id_display'].upper())
            c.fill = PatternFill("solid", fgColor="FFFF99")
            c.alignment = Alignment(vertical='top', wrapText=True)
            c.border = Border(left=Side(border_style="thin", color="444444"),
                              right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Title
            c = ws.cell(row=row, column=2, value=control.get_flattened_oscal_control_as_dict()['title'])
            c.fill = PatternFill("solid", fgColor="FFFF99")
            c.alignment = Alignment(vertical='top', wrapText=True)
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Private Implementation
            smt_combined = ""
            if control.impl_smts:
                for smt in control.impl_smts:
                    smt_combined += smt.body
            c = ws.cell(row=row, column=3, value=smt_combined)
            c.alignment = Alignment(vertical='top', wrapText=True)
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Public Implementation
            c.alignment = Alignment(vertical='top', wrapText=True)
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Notes
            c = ws.cell(row=1, column=5, value="Notes")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Status ["Implemented", "Planned"]
            c = ws.cell(row=1, column=6, value="Status")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Expected Completion (expected implementation)
            c = ws.cell(row=1, column=7, value="Expected Completion")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Class ["Management", "Operational", "Technical",
            c = ws.cell(row=1, column=8, value="Class")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Priority ["p0", "P1", "P2", "P3"]
            c = ws.cell(row=1, column=9, value="Priority")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Responsible Entities
            c = ws.cell(row=1, column=10, value="Responsible Entities")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Control Owner(s)
            c = ws.cell(row=1, column=11, value="Control Owner(s)")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Type ["System-Specific", "Hybrid", "Inherited", "Common", "blank"]
            c = ws.cell(row=1, column=12, value="Type")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Inherited From
            c = ws.cell(row=1, column=13, value="Inherited From")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Provide As ["Do Not Share", "blank"]
            c = ws.cell(row=1, column=14, value="Provide As")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Evaluation Status ["Evaluated", "Expired", "Not Evaluated", "Unknown", "blank"]
            c = ws.cell(row=1, column=15, value="Evaluation Status")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # Control Origination
            c = ws.cell(row=1, column=16, value="Control Origination")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

            # History
            c = ws.cell(row=1, column=17, value="History")
            c.border = Border(right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            blob = stream

        mime_type = "application/octet-stream"
        filename = "{}_control_implementations-{}.xlsx".format(system.root_element.name.replace(" ", "_"),
                                                               datetime.now().strftime("%Y-%m-%d-%H-%M"))

        resp = HttpResponse(blob, mime_type)
        resp['Content-Disposition'] = 'inline; filename=' + filename
        return resp
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def editor(request, system_id, catalog_key, cl_id):
    """System Control detail view"""

    catalog_key, system = get_editor_system(catalog_key, system_id)

    # Retrieve related statements if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        project, catalog, cg_flat, impl_smts, impl_smts_legacy = get_editor_data(request, system, catalog_key, cl_id)
        # TODO: Update system-security-plan to oscal 1.0.0
        # need parties and roles to not be empty
        # Build OSCAL SSP
        # Example: https://github.com/usnistgov/oscal-content/tree/master/examples/ssp/json/ssp-example.json
        # oscalize key
        cl_id = oscalize_control_id(cl_id)

        # Determine previous and next selected controls
        selected_controls_info = list(system.root_element.controls.all().values('id','oscal_ctl_id','oscal_catalog_key'))
        cur_selected_control_index = next((i for i, item in enumerate(selected_controls_info) if (item['oscal_ctl_id'] == cl_id and item['oscal_catalog_key'] == catalog_key)), None)
        if cur_selected_control_index is not None and cur_selected_control_index < len(selected_controls_info)-1:
            next_selected_cl_id = selected_controls_info[cur_selected_control_index + 1]['oscal_ctl_id']
            next_selected_cl_catalog_key = selected_controls_info[cur_selected_control_index + 1]['oscal_catalog_key']
        elif len(selected_controls_info) > 0:
            next_selected_cl_id = selected_controls_info[0]['oscal_ctl_id']
            next_selected_cl_catalog_key = selected_controls_info[0]['oscal_catalog_key']
        else:
            next_selected_cl_id = None
            next_selected_cl_catalog_key = None
        if cur_selected_control_index is not None and cur_selected_control_index > 0:
            prev_selected_cl_id = selected_controls_info[cur_selected_control_index - 1]['oscal_ctl_id']
            prev_selected_cl_catalog_key = selected_controls_info[cur_selected_control_index - 1]['oscal_catalog_key']
        elif len(selected_controls_info) > 1:
            prev_selected_cl_id = selected_controls_info[len(selected_controls_info)-1]['oscal_ctl_id']
            prev_selected_cl_catalog_key = selected_controls_info[len(selected_controls_info)-1]['oscal_catalog_key']
        else:
            prev_selected_cl_id = None
            prev_selected_cl_catalog_key = None

        # Build combined statement if it exists
        if cl_id in system.control_implementation_as_dict:
            combined_smt = system.control_implementation_as_dict[cl_id]['combined_smt']
        else:
            combined_smt = ""

        # Define status options
        impl_statuses = ["Not implemented", "Planned", "Partially implemented", "Implemented", "Unknown"]

      # Only elements for the given control id, sid, and statement type

        elements =  Element.objects.all().exclude(element_type='system')

        context = {
            "system": system,
            "project": project,
            "catalog": catalog,
            "control": cg_flat[cl_id.lower()],
            "impl_smts": impl_smts,
            "impl_statuses": impl_statuses,
            "impl_smts_legacy": impl_smts_legacy,
            "combined_smt": combined_smt,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "opencontrol": "opencontrol_string",
            "elements": elements,
            "display_urls": project_context(project),
            "prev_selected_cl_id": prev_selected_cl_id,
            "prev_selected_cl_catalog_key": prev_selected_cl_catalog_key,
            "next_selected_cl_id": next_selected_cl_id,
            "next_selected_cl_catalog_key": next_selected_cl_catalog_key
        }
        return render(request, "controls/editor.html", context)
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def get_editor_data(request, system, catalog_key, cl_id):
    """
    Get data for editor views
    """

    # Retrieve related statements if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.first()
        parameter_values = project.get_parameter_values(catalog_key)
        catalog = Catalog(catalog_key, parameter_values=parameter_values)
        cg_flat = catalog.get_flattened_controls_all_as_dict()
        # If control id does not exist in catalog
        if cl_id.lower() not in cg_flat:
            return render(request, "controls/detail.html", {"catalog": catalog, "control": {}})

        # Get and return the control
        # Retrieve any related Implementation Statements filtering by control, and system.root_element, Catalog, Type
        impl_smts = Statement.objects.filter(sid=cl_id, consumer_element=system.root_element, sid_class=catalog_key, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name).order_by('pid')
        # Retrieve Legacy Implementation Statements
        impl_smts_legacy = Statement.objects.filter(sid=cl_id, consumer_element=system.root_element, sid_class=catalog_key, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_LEGACY.name)

        return project, catalog, cg_flat, impl_smts, impl_smts_legacy

def get_editor_system(catalog_key, system_id):
    """
    Retrieves oscalized control id and catalog key. Also system object from system id.
    """

    catalog_key = oscalize_catalog_key(catalog_key)

    # Retrieve identified System
    system = System.objects.get(id=system_id)

    return catalog_key, system

@login_required
def editor_compare(request, system_id, catalog_key, cl_id):
    """System Control detail view"""

    catalog_key, system = get_editor_system(catalog_key, system_id)
    # oscalize key
    cl_id = oscalize_control_id(cl_id)
    # Retrieve related statements if owner has permission on system
    if request.user.has_perm('view_system', system):
        project, catalog, cg_flat, impl_smts = get_editor_data(request, system, catalog_key, cl_id)
        context = {
            "system": system,
            "project": project,
            "catalog": catalog,
            "control": cg_flat[cl_id.lower()],
            "impl_smts": impl_smts,
        }
        return render(request, "controls/compare.html", context)
    else:
        # User does not have permission to this system
        raise Http404


# @task_view
@login_required
def save_smt(request):
    """Save a statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # EXAMPLE CODE FOR GUARDIAN PERMISSIONS
        # does user have write privs?
        # if not task.has_write_priv(request.user):
        #     return HttpResponseForbidden()

        # validate question
        # q = task.module.questions.get(id=request.POST.get("question"))

        # validate and parse value
        # if request.POST.get("method") == "clear":
        #     # Clear means that the question returns to an unanswered state.
        #     # This method is only offered during debugging to make it easier
        #     # to test the application's behavior when questions are unanswered.
        #     value = None
        #     cleared = True
        #     skipped_reason = None
        #     unsure = False

        # elif request.POST.get("method") == "skip":
        #     # The question is being skipped, i.e. answered with a null value,
        #     # because the user doesn't know the answer, it doesn't apply to
        #     # the user's circumstances, or they want to return to it later.
        #     value = None
        #     cleared = False
        #     skipped_reason = request.POST.get("skipped_reason") or None
        #     unsure = bool(request.POST.get("unsure"))
        # Track if we are creating a new statement
        new_statement = False
        form_dict = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values['smt_id']
        # Updating or saving a new statement?
        if len(smt_id) > 0:
            # Look up existing Statement object
            statement = Statement.objects.get(pk=smt_id)
            # Check if statement has the same sid class as the statement object
            if statement.sid_class == form_values['sid_class']:
                # Check user permissions
                system_element = statement.consumer_element
                if not request.user.has_perm("change_element", system_element):
                    # User does not have write permissions
                    # Log permission to save answer denied
                    logger.info(
                        event="save_smt permission_denied",
                        object={"object": "statement", "id": statement.id},
                        user={"id": request.user.id, "username": request.user.username}
                    )
                    statement_status = "error"
                    statement_msg = f"Permission denied. {request.user.username} does not have change privileges to system and/or project."
                    return JsonResponse({"status": statement_status, "message": statement_msg})

                if statement is None:
                    # Statement from received has an id no longer in the database.
                    # Report error. Alternatively, in future save as new Statement object
                    statement_status = "error"
                    statement_msg = "The id for this statement is no longer valid in the database."
                    return JsonResponse({"status": statement_status, "message": statement_msg})
                # Update existing Statement object with received info
                statement.pid = form_values['pid']
                statement.body = form_values['body']
                statement.remarks = form_values['remarks']
                statement.status = form_values['status']
            else:
                new_statement = True
        else:
            new_statement_type_enum = StatementTypeEnum[form_values['statement_type'].upper()]
            # Create new Statement object
            new_sid_class = form_values['sid_class'].replace(" ","_") # convert displayed catalog name to catalog_key
            statement = Statement(
                sid=oscalize_control_id(form_values['sid']),
                sid_class=new_sid_class,
                source=new_sid_class,
                body=form_values['body'],
                pid=form_values['pid'],
                statement_type=new_statement_type_enum.name,
                status=form_values['status'],
                remarks=form_values['remarks'],
            )
            new_statement = True

        # handle control_implementation_legacy
        if form_values['statement_type'].upper() == "CONTROL_IMPLEMENTATION_LEGACY":
            # we only need to save control implementation statement
            statement.consumer_element = System.objects.get(pk=form_values['system_id']).root_element
            statement.producer_element = statement.consumer_element
            statement_msg = "Statement associated with System/Consumer Element."
            statement.save()
            statement_element_status = "ok"
            statement_element_msg = "Statement associated with Producer Element."
            messages.add_message(request, messages.INFO, f"{statement_element_msg} {statement.producer_element.id}.")
            # TODO: log result
            serialized_obj = serializers.serialize('json', [statement, ])
            return JsonResponse({"status": "success", "message": statement_msg, "statement": serialized_obj})

        # Updating or saving a new producer_element?
        try:
            # Does the name match and existing element? (Element names are unique.)
            # TODO: Sanitize data entered in form?
            producer_element, created = Element.objects.get_or_create(name=form_values['producer_element_name'])
            if created:
                producer_element.element_type = "system_element"
                producer_element.save()
            producer_element_status = "ok"
            producer_element_msg = "Producer Element saved."
        except Exception as e:
            producer_element_status = "error"
            producer_element_msg = "Producer Element save failed. Error reported {}".format(e)
            return JsonResponse({"status": producer_element_status, "message": producer_element_msg})

        # Associate Statement, Producer Element, and optionally Consumer Element (system) if creating new statement
        if new_statement:
            try:
                statement.producer_element = producer_element
                if 'system_id' in form_values and len(form_values['system_id']) > 0:
                    # Associate Consumer Element
                    statement.consumer_element = System.objects.get(pk=form_values['system_id']).root_element
                    statement_msg = "Statement associated with System/Consumer Element."
                statement.save()
                statement_element_status = "ok"
                statement_element_msg = "Statement associated with Producer Element."
                messages.add_message(request, messages.INFO, f"{statement_element_msg} {producer_element.id}.")
            except Exception as e:
                statement_element_status = "error"
                statement_element_msg = "Failed to associate statement with Producer Element. Error reported {}".format(e)
                return JsonResponse(
                    {"status": statement_element_status, "message": statement_element_msg + " " + producer_element_msg + " " + statement_element_msg})
        # Create new Prototype Statement object on new statement creation (not statement edit)
        if new_statement:
            try:
                statement_prototype = statement.create_prototype()
            except Exception as e:
                statement_status = "error"
                statement_msg = "Statement save failed while saving statement prototype. Error reported {}".format(e)
                return JsonResponse({"status": statement_status, "message": statement_msg})
        # Retain only prototype statement if statement is created in the component library
        # A statement of type `control_implementation` should only exists if associated a consumer_element.
        # When the statement is created in the component library, no consuming_element will exist.
        # TODO
        # - Delete the statement that created the statement prototype
        # - Skip the associating the statement with the system's root_element because we do not have a system identified
        statement_del_msg = ""
        if "form_source" in form_values and form_values['form_source'] == 'component_library':
            # Form received from component library
            serialized_obj = serializers.serialize('json', [statement, ])
            # Delete statement
            Statement.objects.filter(pk=statement.id).delete()
            statement.delete()
            statement_del_msg = "Orphaned Control_Implementation Statement deleted."
        else:
            # Associate Statement and System's root_element
            system_id = form_values['system_id']
            if new_statement and system_id is not None:
                try:
                    statement.consumer_element = System.objects.get(pk=form_values['system_id']).root_element
                    #statement.save()
                    statement_msg = "Statement associated with System/Consumer Element."
                except Exception as e:
                    statement_consumer_status = "error"
                    statement_consumer_msg = "Failed to associate statement with System/Consumer Element {}".format(e)
                    return JsonResponse(
                        {"status": statement_consumer_status, "message": statement_msg + " " + producer_element_msg + " " + statement_consumer_msg})

            # Serialize saved data object(s) to send back to update web page
            # The submitted form needs to be updated with the object primary keys (ids)
            # in order that future saves will be treated as updates.
            serialized_obj = serializers.serialize('json', [statement, ])

    # Save Statement object
    try:
        if not new_statement:
            statement.save()
        statement_msg = "Statement saved."
        messages.add_message(request, messages.INFO, f"Statement {smt_id} Saved")
    except Exception as e:
        statement_status = "error"
        statement_msg = "Statement save failed. Error reported {}".format(e)
        return JsonResponse({"status": statement_status, "message": statement_msg})
    # Return successful save result to web page's Ajax request
    return JsonResponse(
        {"status": "success", "message": statement_msg + " " + producer_element_msg + " " + statement_del_msg,
         "statement": serialized_obj})

@login_required
def update_smt_prototype(request):
    """Update a certified statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # Get statement if exists else 404
        form_dict   = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values['smt_id']
        statement = get_object_or_404(Statement, pk=smt_id)

        # Check permission
        raise_404_if_not_permitted_to_statement(request, statement)

        if statement is None:
            statement_msg = "The id for this statement is no longer valid in the database."
            return JsonResponse({ "status": "error", "message": statement_msg })

        # needs self.body == self.prototype.body

        try:
            proto_statement = Statement.objects.get(pk=statement.prototype_id)
            proto_statement.prototype.body = statement.body
            proto_statement.prototype.save()
            statement_status = "ok"
            statement_msg = f"Update to statement prototype {proto_statement.prototype_id} succeeded."
        except Exception as e:
            statement_msg = "Update to statement prototype failed. Error reported {}".format(e)
            return JsonResponse({ "status": "error", "message": statement_msg })

        return JsonResponse({ "status": "success", "message": statement_msg, "data": { "smt_body": statement.body } })

@login_required
def delete_smt(request):
    """Delete a statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # check permissions
        # EXAMPLE CODE FOR GUARDIAN PERMISSIONS
        # # does user have write privs?
        # # if not task.has_write_priv(request.user):
        # #     return HttpResponseForbidden()

        # # validate question
        # # q = task.module.questions.get(id=request.POST.get("question"))

        # # validate and parse value
        # # if request.POST.get("method") == "clear":
        # #     # Clear means that the question returns to an unanswered state.
        # #     # This method is only offered during debugging to make it easier
        # #     # to test the application's behavior when questions are unanswered.
        # #     value = None
        # #     cleared = True
        # #     skipped_reason = None
        # #     unsure = False

        # # elif request.POST.get("method") == "skip":
        # #     # The question is being skipped, i.e. answered with a null value,
        # #     # because the user doesn't know the answer, it doesn't apply to
        # #     # the user's circumstances, or they want to return to it later.
        # #     value = None
        # #     cleared = False
        # #     skipped_reason = request.POST.get("skipped_reason") or None
        #     unsure = bool(request.POST.get("unsure"))

        form_dict = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values['smt_id']
        # Delete statement?
        statement = Statement.objects.get(pk=smt_id)

        # Check user permissions
        system = statement.consumer_element
        if not request.user.has_perm('change_system', system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="delete_smt permission_denied",
                object={"object": "statement", "id": statement.id},
                user={"id": request.user.id, "username": request.user.username}
            )
            return HttpResponseForbidden(
                "Permission denied. {} does not have change privileges to system and/or project.".format(
                    request.user.username))

        if statement is None:
            # Statement from received has an id no longer in the database.
            # Report error. Alternatively, in future save as new Statement object
            statement_status = "error"
            statement_msg = "The id for this statement is no longer valid in the database."
            return JsonResponse({ "status": "error", "message": statement_msg })
        # Delete Statement object
        try:
            statement.delete()
            statement_status = "ok"
            statement_msg = "Statement deleted."
        except Exception as e:
            statement_status = "error"
            statement_msg = "Statement delete failed. Error reported {}".format(e)
            return JsonResponse({"status": "error", "message": statement_msg})

        # TODO Record fact statement deleted
        # Below will not work because statement is deleted
        # and need to show in record that a statement was recently deleted
        # Update ElementControl smts_updated to know when control element on system was recently updated
        # try:
        #     print("Updating ElementControl smts_updated")
        #     ec = ElementControl.objects.get(element=statement.consumer_element, oscal_ctl_id=statement.sid, oscal_catalog_key=statement.sid_class)
        #     ec.smts_updated = statement.updated
        #     ec.save()
        # except Exception as e:
        #     statement_element_status = "error"
        #     statement_element_msg = "Failed to update ControlElement smt_updated {}".format(e)
        #     return JsonResponse({ "status": "error", "message": statement_msg + " " + producer_element_msg + " " +statement_element_msg })

        return JsonResponse({"status": "success", "message": statement_msg})


# Components
def proposal_message(request, system_id):
    """ Send a global message to indicate a request has been successful """ 
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])


    
    form_dict = dict(request.POST)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]
    messageType = form_values['proposal_message_type']
    message = form_values['proposal_message']
    
    # messageType = request.POST.get("proposal_message_type")
    # message = request.POST.get("proposal_message")

    if(messageType == "INFO"):
        messages.add_message(request, messages.INFO, f'{message}')
    elif(messageType == "WARNING"):
        messages.add_message(request, messages.WARNING, f'{message}')
    else:
        messages.add_message(request, messages.ERROR, f'{message}')

    return HttpResponseRedirect("/systems/{}/components/selected".format(system_id))

def request_message(request, system_id, element_id):
    """ Send a global message to indicate a request has been successful """ 
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    form_dict = dict(request.POST)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]
    messageType = form_values['req_message_type']
    message = form_values['req_message']
    
    if(messageType == "INFO"):
        messages.add_message(request, messages.INFO, f'{message}')
    elif(messageType == "WARNING"):
        messages.add_message(request, messages.WARNING, f'{message}')
    else:
        messages.add_message(request, messages.ERROR, f'{message}')
    return HttpResponseRedirect("/controls/{}/component/{}".format(system_id, element_id))

@login_required
def add_system_component(request, system_id):
    """Add an existing element and its statements to a system"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    form_dict = dict(request.POST)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]

    
    #extract producer_elmentid and require_approval boolean val
    form_values['producer_element_id'], check_req_approval = form_values['producer_element_id'].split(',')
    
    # Does user have permission to add element?
    # Check user permissions
    system = System.objects.get(pk=system_id)
    
    if not request.user.has_perm('change_system', system):
        # User does not have write permissions
        # Log permission to save answer denied
        logger.info(
            event="change_system permission_denied",
            object={"object": "element", "producer_element_name": form_values['producer_element_name']},
            user={"id": request.user.id, "username": request.user.username}
        )
        return HttpResponseForbidden("Permission denied. {} does not have change privileges to system and/or project.".format(request.user.username))

    # DEBUG
    # print(f"Atempting to add {producer_element.name} (id:{producer_element.id}) to system_id {system_id}")

    # Get system's existing components selected
    elements_selected = system.producer_elements
    elements_selected_ids = [e.id for e in elements_selected]

    # Add element to system's selected components
    # Look up the element rto add
    producer_element = Element.objects.get(pk=form_values['producer_element_id'])

    

    # TODO: various use cases
        # - component previously added but element has statements not yet added to system
        #   this issue may be best addressed elsewhere.

    # Component already added to system. Do not add the component (element) to the system again.

    if producer_element.id in elements_selected_ids:
        messages.add_message(request, messages.ERROR,
                            f'Component "{producer_element.name}" already exists in selected components.')
        # Redirect to selected element page
        if "redirect_url" in form_values:
            return HttpResponseRedirect(form_values['redirect_url'])
        else:
            return HttpResponseRedirect("/systems/{}/components/selected".format(system_id))
            
    smts = Statement.objects.filter(producer_element_id = producer_element.id, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name)

    # Component does not have any statements of type control_implementation_prototype to
    # add to system. So we cannot add the component (element) to the system.
    if len(smts) == 0:
        # print(f"The component {producer_element.name} does not have any control implementation statements.")
        messages.add_message(request, messages.ERROR,
                            f'I couldn\'t add "{producer_element.name}" to the system because the component does not currently have any control implementation statements to add.')
        # Redirect to selected element page
        if "redirect_url" in form_values:
            return HttpResponseRedirect(form_values['redirect_url'])
        else:
            return HttpResponseRedirect("/systems/{}/components/selected".format(system_id))

    # Loop through all element's prototype statements and add to control implementation statements.
    # System's selected controls will filter what controls and control statements to display.
    for smt in smts:
        smt.create_system_control_smt_from_component_prototype_smt(system.root_element.id)

    # Make sure some controls were added to the system. Report error otherwise.
    smts_added = Statement.objects.filter(producer_element_id = producer_element.id, consumer_element_id = system.root_element.id, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name)

    smts_added_count = len(smts_added)
    if smts_added_count > 0:
        messages.add_message(request, messages.INFO,
                         f'OK. I\'ve added "{producer_element.name}" to the system and its {smts_added_count} control implementation statements to the system.')
    else:
        messages.add_message(request, messages.WARNING,
                         f'Oops. I tried adding "{producer_element.name}" to the system, but the component added 0 controls.')

    # Redirect to selected element page
    if "redirect_url" in form_values:
        return HttpResponseRedirect(form_values['redirect_url'])
    else:
        return HttpResponseRedirect("/systems/{}/components/selected".format(system_id))
  
@login_required
def search_system_component(request):
    """Add an existing element and its statements to a system"""

    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])

    form_dict = dict(request.GET)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]
    # Form values from ajax data

    if "system_id" in form_values.keys():
        system_id = form_values['system_id']

        # Does user have permission to add element?
        # Check user permissions
        system = System.objects.get(pk=system_id)

        if not request.user.has_perm('change_system', system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="change_system permission_denied",
                object={"object": "element", "producer_element_name": form_values['producer_element_name']},
                user={"id": request.user.id, "username": request.user.username}
            )
            return HttpResponseForbidden("Permission denied. {} does not have change privileges to system and/or project.".format(request.user.username))

        selected_controls = system.root_element.controls.all()
        selected_controls_ids = set()
        for sc in selected_controls:
            selected_controls_ids.add("{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key))

        # Add element
        # Look up the element

        text =  form_values['text']

        # The final elements that are returned to the new dropdown created...
        producer_system_elements = Element.objects.filter(element_type="system_element").filter(name__contains= text)

        producer_elements = [{"id":str(ele.id), "name": ele.name} for ele in producer_system_elements]

        results = {'producer_element_statement_values': producer_elements}
        data = json.dumps(results)
        mimetype = 'application/json'
        return HttpResponse(data, mimetype)
    # # Redirect to selected element page
    # return HttpResponseRedirect("/systems/{}/components/selected".format(system_id))

class RelatedComponentStatements(View):
    """
    Returns the component statements that are produced(related) to one control implementation prototype
    """

    template_name = 'controls/editor.html'

    def get(self, request):
        """Add an existing element and its statements to a system"""

        if request.method != "GET":
            return HttpResponseNotAllowed(["GET"])

        logger.info(f"Related controls GET: {dict(request.GET)}")
        form_dict = dict(request.GET)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values['system_id']

            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)

            if not request.user.has_perm('change_system', system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={"object": "element", "producer_element_name": form_values['producer_element_name']},
                    user={"id": request.user.id, "username": request.user.username}
                )
                return HttpResponseForbidden(
                    "Permission denied. {} does not have change privileges to system and/or project.".format(
                        request.user.username))

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add("{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key))

            # Add element
            # Look up the element
            producer_element_id = form_values['producer_element_form_id']
            producer_element = Element.objects.get(pk=producer_element_id)

            # The final elements that are returned to the new dropdown created...
            producer_smt_imps = producer_element.statements("control_implementation")

            producer_smt_imps_vals = [{"smt_id": str(smt.id), "smt_sid": smt.sid, "smt_sid_class": smt.sid_class,  "producer_element_name": producer_element.name, "smt_body":str(smt.body), "smt_pid":str(smt.pid), "smt_status":str(smt.status)} for smt in producer_smt_imps]

            results = {'producer_element_statement_values': producer_smt_imps_vals,  "selected_component": producer_element.name}
            data = json.dumps(results)
            mimetype = 'application/json'
            if data:
                return HttpResponse(data, mimetype)
            else:
                return JsonResponse(status=400, data={'status': 'error', 'message': f"No DATA!: {data}"})
        else:
            return JsonResponse(status=400, data={'status': 'error', 'message': "There is no current system id present"})

class EditorAutocomplete(View):
    template_name = 'controls/editor.html'

    def get(self, request):
        """Add an existing element and its statements to a system"""

        if request.method != "GET":
            return HttpResponseNotAllowed(["GET"])

        form_dict = dict(request.GET)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values['system_id']

            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)

            if not request.user.has_perm('change_system', system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={"object": "element", "producer_element_name": form_values['producer_element_name']},
                    user={"id": request.user.id, "username": request.user.username}
                )
                return HttpResponseForbidden(
                    "Permission denied. {} does not have change privileges to system and/or project.".format(
                        request.user.username))

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add("{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key))

            # Add element
            # Look up the element

            text = form_values['text']

            # The final elements that are returned to the new dropdown created...
            producer_system_elements = Element.objects.filter(element_type="system_element").filter(name__contains=text)

            producer_elements = [{"id": str(ele.id), "name": ele.name} for ele in producer_system_elements]

            results = {'producer_element_name_value': producer_elements}
            data = json.dumps(results)
            mimetype = 'application/json'
            if data:
                return HttpResponse(data, mimetype)
            else:
                return JsonResponse(status=400, data={'status': 'error', 'message': f"No statements found with the search: {text}"})
        else:
            return JsonResponse(status=400, data={'status': 'error', 'message': "There is no current system id present"})

    def post(self, request, system_id):
        """Add an existing element and its statements to a system"""
        if request.method != "POST":
            return HttpResponseNotAllowed(["POST"])

        form_dict = dict(request.POST)

        form_values = {}
        for key in form_dict.keys():
            if key == 'relatedcomps':
                form_values[key] = form_dict[key]
            else:
                form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values['system_id']
            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)
            if not request.user.has_perm('change_system', system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={"object": "element", "entered_producer_element_name": form_values['text'] or "None"},
                    user={"id": request.user.id, "username": request.user.username}
                )
                return HttpResponseForbidden(
                    "Permission denied. {} does not have change privileges to system and/or project.".format(
                        request.user.username))

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add("{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key))

            # Add element
            if form_values.get("relatedcomps", ""):

                for related_element in form_values['relatedcomps']:

                    # Look up the element
                    for smt in Statement.objects.filter(id=related_element, statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name):
                        logger.info(
                            f"Adding an element with the id {smt.id} and sid class {smt.sid} to system_id {system_id}")
                        # Only add statements for controls selected for system
                        if "{} {}".format(smt.sid, smt.sid_class) in selected_controls_ids:
                            logger.info(f"smt {smt}")
                            smt.create_system_control_smt_from_component_prototype_smt(system.root_element.id)
                        else:
                            logger.error(f"not adding smt from selected controls for the current system: {smt}")

        # Redirect to the page where the component was added from
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


# Baselines
@login_required
def assign_baseline(request, system_id, catalog_key, baseline_name):
    """Assign a baseline to a system root element thereby showing selected controls for the system."""

    system = System.objects.get(pk=system_id)
    #system.root_element.assign_baseline_controls(user, 'NIST_SP-800-53_rev4', 'low')
    assign_results = system.root_element.assign_baseline_controls(request.user, catalog_key, baseline_name)
    if assign_results:
        messages.add_message(request, messages.INFO,
                             'Baseline "{} {}" assigned.'.format(catalog_key.replace("_", " "), baseline_name.title()))
        # Log start app / new project
        logger.info(
            event="assign_baseline",
            object={"object": "system", "id": system.root_element.id, "title": system.root_element.name},
            baseline={"catalog_key": catalog_key, "baseline_name": baseline_name},
            user={"id": request.user.id, "username": request.user.username}
        )
    else:
        messages.add_message(request, messages.ERROR,
                             'Baseline "{} {}" assignment failed.'.format(catalog_key.replace("_", " "),
                                                                          baseline_name.title()))

    return HttpResponseRedirect(f"/systems/{system_id}/controls/selected")


# Export OpenControl
@login_required
def export_system_opencontrol(request, system_id):
    """Export entire system in OpenControl"""

    # Does user have permission
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        # Create temporary directory structure
        import tempfile
        temp_dir = tempfile.TemporaryDirectory(dir=".")
        repo_path = os.path.join(temp_dir.name, system.root_element.name.replace(" ", "_"))
        if not os.path.exists(repo_path):
            os.makedirs(repo_path)

        # Create various directories
        os.makedirs(os.path.join(repo_path, "components"))
        os.makedirs(os.path.join(repo_path, "standards"))
        os.makedirs(os.path.join(repo_path, "certifications"))

        # Create opencontrol.yaml config file
        cfg_str = """schema_version: 1.0.0
name: ~
metadata:
  authorization_id: ~
  description: ~
  organization:
    name: ~
    abbreviation: ~
  repository: ~
components: []
standards:
- ./standards/NIST-SP-800-53-rev4.yaml
certifications:
- ./certifications/fisma-low-impact.yaml
"""

        # read default opencontrol.yaml into object
        cfg = rtyaml.load(cfg_str)
        # customize values
        cfg["name"] = system.root_element.name
        # cfg["metadata"]["organization"]["name"] = organization_name
        # cfg["metadata"]["description"] = description
        # cfg["metadata"]["organization"]["abbreviation"] = None
        # if organization_name:
        #     cfg["metadata"]["organization"]["abbreviation"] = "".join([word[0].upper() for word in organization_name.split(" ")])

        with open(os.path.join(repo_path, "opencontrol.yaml"), 'w') as outfile:
            outfile.write(rtyaml.dump(cfg))

        # Populate reference directories from reference
        OPENCONTROL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'opencontrol')
        shutil.copyfile(os.path.join(OPENCONTROL_PATH, "standards", "NIST-SP-800-53-rev4.yaml"),
                        os.path.join(repo_path, "standards", "NIST-SP-800-53-rev4.yaml"))
        shutil.copyfile(os.path.join(OPENCONTROL_PATH, "standards", "NIST-SP-800-171r1.yaml"),
                        os.path.join(repo_path, "standards", "NIST-SP-800-53-rev4.yaml"))
        shutil.copyfile(os.path.join(OPENCONTROL_PATH, "standards", "opencontrol.yaml"),
                        os.path.join(repo_path, "standards", "opencontrol.yaml"))
        shutil.copyfile(os.path.join(OPENCONTROL_PATH, "standards", "hipaa-draft.yaml"),
                        os.path.join(repo_path, "standards", "hipaa-draft.yaml"))
        shutil.copyfile(os.path.join(OPENCONTROL_PATH, "certifications", "fisma-low-impact.yaml"),
                        os.path.join(repo_path, "certifications", "fisma-low-impact.yaml"))

        # # Make stub README.md file
        # with open(os.path.join(repo_path, "README.md"), 'w') as outfile:
        #     outfile.write("Machine readable representation of 800-53 control implementations for {}.\n\n# Notes\n\n".format(system_name))
        #     print("wrote file: {}\n".format(os.path.join(repo_path, "README.md")))

        # Populate system information files

        # Populate component files
        if not os.path.exists(os.path.join(repo_path, "components")):
            os.makedirs(os.path.join(repo_path, "components"))
        for element in system.producer_elements:
            # Build OpenControl
            ocf = {
                "name": element.name,
                "schema_version": "3.0.0",
                "documentation_complete": False,
                "satisfies": []
            }
            satisfies_smts = ocf["satisfies"]
            # Retrieve impl_smts produced by element and consumed by system
            # Get the impl_smts contributed by this component to system
            impl_smts = element.statements_produced.filter(consumer_element=system.root_element)
            for smt in impl_smts:
                my_dict = {
                    "control_key": smt.sid.upper(),
                    "control_name": smt.catalog_control_as_dict['title'],
                    "standard_key": smt.sid_class,
                    "covered_by": [],
                    "security_control_type": "Hybrid | Inherited | ...",
                    "narrative": [
                        {"text": smt.body}
                    ],
                    "remarks": [
                        {"text": smt.remarks}
                    ]
                }
                satisfies_smts.append(my_dict)
            opencontrol_string = rtyaml.dump(ocf)
            # Write component file
            with open(os.path.join(repo_path, "components", "{}.yaml".format(element.name.replace(" ", "_"))), 'w') as fh:
                fh.write(opencontrol_string)

        # Build Zip archive
        # TODO Make Temporary File
        #      Current approach leads to race conditions!
        shutil.make_archive("/tmp/Zipped_file", 'zip', repo_path)

        # Download Zip archive of OpenControl files
        with open('/tmp/Zipped_file.zip', 'rb') as tmp:
            tmp.seek(0)
            stream = tmp.read()
            blob = stream
        mime_type = "application/octet-stream"
        filename = "{}-opencontrol-{}.zip".format(system.root_element.name.replace(" ", "_"),
                                                  datetime.now().strftime("%Y-%m-%d-%H-%M"))

        resp = HttpResponse(blob, mime_type)
        resp['Content-Disposition'] = 'inline; filename=' + filename

        # Clean up
        shutil.rmtree(repo_path)
        # os.remove("/tmp/Zipped_file") ????
        return resp

    else:
        # User does not have permission to this system
        raise Http404

# Poams
@login_required
def system_poams_aspen(request, system_id):
    """System Summary page experiment POA&Ms"""

    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    system_events = get_integrations_system_events(request, system_id)

    # Fix menu data for old vertical menu
    project.system.root_element.name = system_summary['name']
    project.root_task.title_override = system_summary['name']

    controls = system.root_element.controls.all()
    poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('-updated')

    # Example reading POA&Ms from xlsx file using Pandas
    # This is to just demonstrate reading POA&Ms from imported spreadsheet
    # TODO: Move to a serializer and filter for one sysyem only
    # import pandas
    # poams_list = []
    # fn = "local/poams_list.xlsx"
    # if pathlib.Path(fn).is_file():
    #     try:
    #         df_dict = pandas.read_excel(fn, header=1)
    #         for index, row in df_dict.iterrows():
    #             poam_dict = {
    #                 "id": row.get('CSAM ID', ""),
    #                 "CSAM_ID": row.get('CSAM ID', ""   ),
    #                 "Org": row.get('Org', ""   ),
    #                 "Sub_Org": row.get('Sub Org', ""   ),
    #                 "System_Name": row.get('System Name', ""   ),
    #                 "POAM_ID": row.get('POAM ID', ""   ),
    #                 "POAM_Title": row.get('POAM Title', "" ),
    #                 "System_Type": row.get('System Type', ""   ),
    #                 "Detailed_Weakness_Description": row.get('Detailed Weakness Description', ""   ),
    #                 "Status": row.get('Status', "" )
    #             }
    #             poams_list.append(poam_dict)
    #     except FileNotFoundError as e:
    #         logger.error(f"Error reading file {fn}: {e}")
    #     except Exception as e:
    #         logger.error(f"Other Error reading file {fn}: {e}")

    context = {
        "system": system_summary,
        "poam_smts": poam_smts,
        "display_urls": project_context(project)
    }
    return render(request, "systems/poams_list_aspen.html", context)

@login_required
def poams_list(request, system_id):
    """List Poams for a system"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('-updated')

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "controls": controls,
            "poam_smts": poam_smts,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "display_urls": project_context(project)
        }
        return render(request, "systems/poams_list.html", context)
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def new_poam(request, system_id):
    """Form to create new POAM"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()

        if request.method == 'POST':
            statement_form = StatementPoamForm(request.POST)
            # if statement_form.is_valid() and poam_form.is_valid():
            if statement_form.is_valid():
                statement = statement_form.save()
                poam_form = PoamForm(request.POST)
                if poam_form.is_valid():
                    poam = poam_form.save()
                    print('POAM ID', poam.get_next_poam_id(system))
                    poam.poam_id = poam.get_next_poam_id(system)
                    poam.statement = statement
                    poam.save()
                return HttpResponseRedirect('/systems/{}/poams'.format(system_id), {})
                #url(r'^(?P<system_id>.*)/poams$', views.poams_list, name="poams_list"),
            else:
                pass
        else:
            statement_form = StatementPoamForm(status="Open", statement_type="POAM", consumer_element=system.root_element)
            poam = Poam()
            poam_id = poam.get_next_poam_id(system)
            poam_form = PoamForm()
            return render(request, 'systems/poam_form.html', {
                'statement_form': statement_form,
                'poam_form': poam_form,
                'system': system,
                'project': project,
                'controls': controls,
                "display_urls": project_context(project)
            })
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def edit_poam(request, system_id, poam_id):
    """Form to create new POAM"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        # Retrieve POAM Statement
        poam_smt = get_object_or_404(Statement, id=poam_id)

        if request.method == 'POST':
            statement_form = StatementPoamForm(request.POST, instance=poam_smt)
            poam_form = PoamForm(request.POST, instance=poam_smt.poam)
            if statement_form.is_valid() and poam_form.is_valid():
                # Save statement after updating values
                statement_form.save()
                poam_form.save()
                return HttpResponseRedirect('/systems/{}/poams'.format(system_id), {})
            else:
                pass
                #TODO: What if form invalid?
        else:
            statement_form = StatementPoamForm(initial={
                'statement_type': poam_smt.statement_type,
                'status': poam_smt.status,
                'consumer_element': system.root_element,
                'body': poam_smt.body,
                'remarks': poam_smt.remarks,
            })
            poam_form = PoamForm(initial={
                'weakness_name': poam_smt.poam.weakness_name,
                'controls': poam_smt.poam.controls,
                'poam_group': poam_smt.poam.poam_group,
                'risk_rating_original': poam_smt.poam.risk_rating_original,
                'risk_rating_adjusted': poam_smt.poam.risk_rating_adjusted,
                'weakness_detection_source': poam_smt.poam.weakness_detection_source,
                'remediation_plan': poam_smt.poam.remediation_plan,
                'milestones': poam_smt.poam.milestones,
                'scheduled_completion_date': poam_smt.poam.scheduled_completion_date,
            })
            return render(request, 'systems/poam_edit_form.html', {
                'statement_form': statement_form,
                'poam_form': poam_form,
                'system': system,
                'project': project,
                'controls': controls,
                'poam_smt': poam_smt,
                "display_urls": project_context(project)
            })
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def poam_export_xlsx(request, system_id):
    return poam_export(request, system_id, 'xlsx')

@login_required
def poam_export_csv(request, system_id):
    return poam_export(request, system_id, 'csv')

@login_required
def poam_export(request, system_id, format='xlsx'):
    """Export POA&M in either xlsx or csv"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected POA&Ms if user has permission on system
    if request.user.has_perm('view_system', system):

        if format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
            from tempfile import NamedTemporaryFile

            wb = Workbook()
            ws = wb.active
            # create alignment style
            wrap_alignment = Alignment(wrap_text=True)
            ws.title = "POA&Ms"
        else:
            import csv, io
            csv_buffer = io.StringIO(newline='\n')
            csv_writer = csv.writer(csv_buffer)

        poam_fields = [
            {'var_name': 'poam_id', 'name': 'POA&M ID', 'width': 8},
            {'var_name': 'poam_group', 'name': 'POA&M Group', 'width': 16},
            {'var_name': 'weakness_name', 'name': 'Weakness Name', 'width': 24},
            {'var_name': 'controls', 'name': 'Controls', 'width': 16},
            {'var_name': 'body', 'name': 'Description', 'width': 60},
            {'var_name': 'status', 'name': 'Status', 'width': 8},
            {'var_name': 'risk_rating_original', 'name': 'Risk Rating Original', 'width': 16},
            {'var_name': 'risk_rating_adjusted', 'name': 'Risk Rating Adjusted', 'width': 16},
            {'var_name': 'weakness_detection_source', 'name': 'Weakness Detection Source', 'width': 24},
            {'var_name': 'weakness_source_identifier', 'name': 'Weakness Source Identifier', 'width': 24},
            {'var_name': 'remediation_plan', 'name': 'Remediation Plan', 'width': 60},
            {'var_name': 'milestones', 'name': 'Milestones', 'width': 60},
            {'var_name': 'milestone_changes', 'name': 'Milestone Changes', 'width': 30},
            {'var_name': 'scheduled_completion_date', 'name': 'Scheduled Completion Date', 'width': 18},
        ]

        # create header row
        column = 0
        ord_zeroth_column = ord('A') - 1
        csv_row = []

        for poam_field in poam_fields:
            column += 1
            if format == 'xlsx':
                c = ws.cell(row=1, column=column, value=poam_field['name'])
                c.fill = PatternFill("solid", fgColor="5599FE")
                c.font = Font(color="FFFFFF", bold=True)
                c.border = Border(left=Side(border_style="thin", color="444444"),
                                  right=Side(border_style="thin", color="444444"),
                                  bottom=Side(border_style="thin", color="444444"),
                                  outline=Side(border_style="thin", color="444444"))
                ws.column_dimensions[chr(ord_zeroth_column + column)].width = poam_field['width']
            else:
                csv_row.append(poam_field['name'])
        # Add column for URL
        if format == 'xlsx':
            c = ws.cell(row=1, column=column, value="URL")
            c.fill = PatternFill("solid", fgColor="5599FE")
            c.font = Font(color="FFFFFF", bold=True)
            c.border = Border(left=Side(border_style="thin", color="444444"),
                              right=Side(border_style="thin", color="444444"),
                              bottom=Side(border_style="thin", color="444444"),
                              outline=Side(border_style="thin", color="444444"))
            ws.column_dimensions[chr(ord_zeroth_column + column)].width = 60
        else:
            csv_row.append('URL')

        if format != 'xlsx':
            csv_writer.writerow(csv_row)

        # Retrieve POA&Ms and create POA&M rows
        poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('id')
        poam_smts_by_sid = {}
        row = 1
        for poam_smt in poam_smts:
            csv_row = []
            row += 1

            # Loop through fields
            column = 0
            for poam_field in poam_fields:
                column += 1
                if format == 'xlsx':
                    if poam_field['var_name'] in ['body', 'status']:
                        c = ws.cell(row=row, column=column, value=getattr(poam_smt, poam_field['var_name']))
                    else:
                        if poam_field['var_name'] == 'poam_id':
                            c = ws.cell(row=row, column=column,
                                        value="V-{}".format(getattr(poam_smt.poam, poam_field['var_name'])))
                        else:
                            c = ws.cell(row=row, column=column, value=getattr(poam_smt.poam, poam_field['var_name']))
                    c.fill = PatternFill("solid", fgColor="FFFFFF")
                    c.alignment = Alignment(vertical='top', horizontal='left', wrapText=True)
                    c.border = Border(right=Side(border_style="thin", color="444444"),
                                      bottom=Side(border_style="thin", color="444444"),
                                      outline=Side(border_style="thin", color="444444"))
                else:
                    if poam_field['var_name'] in ['body', 'status']:
                        csv_row.append(getattr(poam_smt, poam_field['var_name']))
                    else:
                        if poam_field['var_name'] == 'poam_id':
                            csv_row.append("V-{}".format(getattr(poam_smt.poam, poam_field['var_name'])))
                        else:
                            csv_row.append(getattr(poam_smt.poam, poam_field['var_name']))

            # Add URL column
            poam_url = settings.SITE_ROOT_URL + "/systems/{}/poams/{}/edit".format(system_id, poam_smt.id)
            if format == 'xlsx':
                c = ws.cell(row=row, column=column, value=poam_url)
                c.fill = PatternFill("solid", fgColor="FFFFFF")
                c.alignment = Alignment(vertical='top', horizontal='left', wrapText=True)
                c.border = Border(right=Side(border_style="thin", color="444444"),
                                  bottom=Side(border_style="thin", color="444444"),
                                  outline=Side(border_style="thin", color="444444"))
            else:
                csv_row.append(poam_url)

            if format != 'xlsx':
                csv_writer.writerow(csv_row)

        if format == 'xlsx':
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
                blob = stream
        else:
            blob = csv_buffer.getvalue()
            csv_buffer.close()

        # Determine filename based on system name
        system_name = system.root_element.name.replace(" ", "_") + "_" + system_id
        filename = "{}_poam_export-{}.{}".format(system_name, datetime.now().strftime("%Y-%m-%d-%H-%M"), format)
        mime_type = "application/octet-stream"

        resp = HttpResponse(blob, mime_type)
        resp['Content-Disposition'] = 'inline; filename=' + filename
        return resp
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def system_csv_template(request):
    import csv
    # Create the HttpResponse object with the appropriate CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=system_template.csv'
    # Create the CSV writer using the HttpResponse as the "file"
    writer = csv.writer(response)
    writer.writerow([
        'System Name', 
        'System Description', 
        'System UUID', 
        'System Private', 
        'System Require Approval', 
        'System Created', 
        'System Updated',
        'System OSCAL Version',
        'System Component Type', 
        'System Component State'
    ])

    return response

@login_required
def system_xlsx_template(request):
    # response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="system_template.xlsx"'
    # df = pandas.DataFrame([],column=['system_name', 'system_description'])
    # print('df: ', df)
    # df.to_excel('system_template.xlsx', index=False, header=True)

    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from tempfile import NamedTemporaryFile

    wb = Workbook()
    ws = wb.active
    # create alignment style
    wrap_alignment = Alignment(wrap_text=True)
    ws.title = "Sheet1"
    system_headers = [
        {"var_name": "csam_id", "name": "CSAM ID", "width": 32},
        {"var_name": "org", "name": "Org", "width": 32},
        {"var_name": "sub_org", "name": "Sub Org", "width": 32},
        {"var_name": "system_name", "name": "System Name", "width": 32},
        {"var_name": "system_description", "name": "System Description", "width": 32},
        {"var_name": "acronym", "name": "Acronym", "width": 32},
        {"var_name": "system_category", "name": "System Category", "width": 32},
        {"var_name": "system_operational_status", "name": "System Operational Status", "width": 32},
        {"var_name": "system_type", "name": "System Type", "width": 32},
        {"var_name": "contractor_system", "name": "Contractor System", "width": 32},
        {"var_name": "financial_system", "name": "Financial System", "width": 32},
        {"var_name": "fisma_reportable", "name": "FISMA Reportable", "width": 32},
        {"var_name": "critical_infrastructure", "name": "Critical Infrastructure", "width": 32},
        {"var_name": "mission_critical", "name": "Mission Critical", "width": 32},
        {"var_name": "uui_code", "name": "UII Code", "width": 32},
        {"var_name": "investment_name", "name": "Investment Name", "width": 32},
        {"var_name": "portfolio", "name": "Portfolio", "width": 32},
        {"var_name": "poam_id", "name": "POAM ID", "width": 32},
        {"var_name": "poam_sequence", "name": "POAM Sequence", "width": 32},
        {"var_name": "poam_title", "name": "POAM Title", "width": 32},
        {"var_name": "detailed_weakness_description", "name": "Detailed Weakness Description", "width": 32},
        {"var_name": "create_date", "name": "Create Date", "width": 32},
        {"var_name": "days_since_creation", "name": "Days Since Creation", "width": 32},
        {"var_name": "scheduled_completion_date", "name": "Scheduled Completion Date", "width": 32},
        {"var_name": "planned_start_date", "name": "Planned Start Date", "width": 32},
        {"var_name": "planned_finish_date", "name": "Planned Finish Date", "width": 32},
        {"var_name": "planned_finish_date", "name": "Planned Finish Date", "width": 32},
        {"var_name": "status", "name": "Status", "width": 32},
        {"var_name": "weakness", "name": "Weakness", "width": 32},
        {"var_name": "cost", "name": "Cost", "width": 32},
        {"var_name": "control_risk_severity", "name": "Control Risk Severity", "width": 32},
        {"var_name": "user_identified_critically", "name": "User Identified Critically", "width": 32},
        {"var_name": "severity", "name": "Severity", "width": 32},
        {"var_name": "workflow_status", "name": "Workflow Status", "width": 32},
        {"var_name": "workflow_status_date", "name": "Workflow Status Date", "width": 32},
        {"var_name": "days_until_auto_approved", "name": "Days Until Auto-Approved", "width": 32},
        {"var_name": "exclude_from_omb", "name": "Exclude From OMB", "width": 32},
        {"var_name": "accepted_risk", "name": "Accepted Risk", "width": 32},
        {"var_name": "assigned_to", "name": "Assigned To", "width": 32},
        {"var_name": "phone", "name": "Phone", "width": 32},
        {"var_name": "email", "name": "Email", "width": 32},
        {"var_name": "assigned_date", "name": "Assigned Date", "width": 32},
        {"var_name": "delay_reason", "name": "Delay Reason", "width": 32},
        {"var_name": "controls", "name": "Controls", "width": 32},
        {"var_name": "csffunction", "name": "CSFFunction", "width": 32},
        {"var_name": "csfcategory", "name": "CSFCategory", "width": 32},
        {"var_name": "csfsubcategory", "name": "CSFSubCategory", "width": 32},
        {"var_name": "number_milestones", "name": "Number Milestones", "width": 32},
        {"var_name": "number_artifacts", "name": "Number Artifacts", "width": 32},
        {"var_name": "rbd_approval_date", "name": "RBD Approval Date", "width": 32},
        {"var_name": "deficiency_category", "name": "Deficiency Category", "width": 32},
        {"var_name": "source_of_finding", "name": "Source of Finding", "width": 32},
        {"var_name": "percent_complete", "name": "Percent Complete", "width": 32},
        {"var_name": "date_percent_complete_late_updated", "name": "Date % Complete Last Updated", "width": 32}
    ]

     # create header row
    column = 0
    ord_zeroth_column = ord('A') - 1

    for header in system_headers:
        column += 1
        c = ws.cell(row=1, column=column, value=header['name'])
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(left=Side(border_style="thin", color="444444"),
                            right=Side(border_style="thin", color="444444"),
                            bottom=Side(border_style="thin", color="444444"),
                            outline=Side(border_style="thin", color="444444"))
        # ws.column_dimensions[chr(ord_zeroth_column + column)].width = header['width']

    filename = "system_template.xlsx"
    mime_type = "application/octet-stream"
    with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
                blob = stream
    response = HttpResponse(blob, mime_type)
    response['Content-Disposition'] = 'inline; filename=' + filename
    return response

@login_required
def system_oscal_template(request):
    system_example = {
        "uuid": str(uuid.uuid4()),
        "name": "New System Name",
        "description": "New System Description"
    }
    
    of = {
        "system-security-plan": {
            "uuid": system_example["uuid"],
            "metadata": {
                "title": "{} System Security Plan".format(system_example["name"]),
                "last-modified": datetime.today().replace(microsecond=0).isoformat(),
                "version": "1.0",
                "oscal-version": "1.0.0",
                "roles": [
                    {
                        "id": "admin", 
                        "title": "admin"
                    }
                ],
                "parties": [
                    {
                        "uuid": str(uuid.uuid4()),
                        "type": "organization",
                        "name": "organization name"
                    }
                ],
            },
            "import-profile": {
                "href": "http://localhost:8000/profile_path"
            },
            "system-characteristics": {
                "system-name": system_example["name"],
                "description": system_example["description"],
                "system-ids":[
                    {
                        "id": system_example["uuid"],
                        "identifier-type": "https://ietf.org/rfc/rfc4122"
                    }
                ]
            },
            "system-implementation": {
                "remarks": "",
                "users": [],
                "components": [],
                "inventory-items": [        
                    {
                        "uuid": str(uuid.uuid4()),
                        "description": "An inventory item",
                        "implemented-components": [
                            {
                                "uuid": str(uuid.uuid4()),
                                "description": "An inventory item",
                                "implemented-components": [
                                    {
                                    "component-uuid": str(uuid.uuid4())
                                    }
                                ]
                            }
                        ]
                    }
                ]
            },
            "control-implementation": {
                "description": "",
                "implemented-requirements": [], 
            }
        }
    }

    try:
        # Create a temporary directory and dump the json_object in there.
        tempdir = tempfile.mkdtemp()
        path = os.path.join(tempdir, "ssp_object.json")
        # Use trestle's ComponentDefinition method oscal_read to read the path to json in the temporary folder
        path_ssp_definition = pathlib.Path(path)

        with open(path, 'w+') as cred:
            json.dump(of, cred)
        # Read in temporary file and shape into trestle pydantic SSP definition.
        trestle_oscal_json = trestlessp.SystemSecurityPlan.oscal_read(path_ssp_definition)
        # Finally validate that this object is valid by the OSCAL System Security Plan definition
        trestlessp.SystemSecurityPlan.validate(trestle_oscal_json)
        # Cleanup
        shutil.rmtree(tempdir)
    except Exception as e:
        logger.error(f"Invalid System Security Plan JSON: {e}")
        shutil.rmtree(tempdir)
        return HttpResponse(e)

    oscal_string = json.dumps(of, sort_keys=False, indent=2)
    return oscal_string

import uuid

def is_valid_uuid(val):
    try:
        uuid.UUID(str(val))
        return True
    except ValueError:
        return False

def validate_field(validation_type, value):
    """
    Evaluates a value against a known value type, and returns desired evaluated value
    """
    if validation_type == "System Name":
        if value:
            return value
        else:
            return {
                validation_type: "Missing System Name",
            }
    if validation_type == "System Description":
        if value:
            return value
        else:
            return {
                validation_type: "Missing System Description",
            }
    if validation_type == "System UUID":
        if is_valid_uuid(value):
            return value
        else:
            return {
                validation_type: "Invalid UUID format",
            }
    if validation_type == "System Private":
        if value == True or value == False:
            return value
        else:
            return None
    if validation_type == "System Require Approval":
        if value == True or value == False:
            return value
        else:
            return {
                validation_type: "Invalid boolean value: please use True or False",
            }
    if validation_type == "System Created":
        # dateFormat = re.search(r'^[0-9]{1,2}\/[0-9]{1,2}\/[0-9]{}$', value)
        try:
            created = datetime.strptime(value, '%m/%d/%y').replace(tzinfo = timezone.utc)
            return created
        except:
            return None
    if validation_type == "System Updated":
        try: 
            updated = datetime.strptime(value, '%m/%d/%y').replace(tzinfo = timezone.utc)
            return updated
        except:
            return {
                validation_type: "Invalid Date format, please use mm/dd/yyyy",
            }
    if "date" in validation_type.lower():
        if value:
            try:
                date = datetime.strptime(value, '%m/%d/%y').replace(tzinfo = timezone.utc)
                return date
            except:
                try:
                    date = datetime.strptime(value, '%m/%d/%Y').replace(tzinfo = timezone.utc)
                    return date
                except:
                    try:
                        date = datetime.strptime(value, '%m-%d-%Y %H:%M:%S %p').replace(tzinfo = timezone.utc)
                        return date
                    except:
                        return {
                            validation_type: "Invalid Date format, please use mm/dd/yyyy",
                        }
        else:
            return None
    if validation_type == "System OSCAL Version":
        isOSCALVersionFormatValid = re.search(r'^[0-9]+\.[0-9]*\.[0-9]*[-[a-zA-Z0-9]*]?$', str(value))
        if isOSCALVersionFormatValid:
            return value
        else:
            return {
                validation_type: "Invalid OSCAL version format, please use #.#.# format, optional #.#.#-L#",
            }
    if validation_type == "System Component Type":
         # Check if input is one of the accepted values - []
        validTypes = ["hardware", "software", "service", "policy", "process", "procedure"]
        if value.lower() in validTypes:
            return value
        else:
            return {
                validation_type: "Invalid System component type: please use one of the following options: hardware, software, service, policy, process, procedure",
            }
    if validation_type == "System Component State":
         # Check if input is one of the accepted values - []
        validStates = ["under-development", "operational", "disposition", "other"]
        if value.lower() in validStates:
            return value
        else:
            return {
                validation_type: "Invalid System component state: please use one of the following options: under-development, operational, disposition, other",
            }
    return value

# Project
@login_required
def system_import(request):
    """
    Import a system information and create a new system
    """

    # Retrieve identified System
    if request.method == 'POST':
        
        file_content = request.POST['file_content']
        file_name = request.POST['file_name']

        systems_imported_list = set()
        systems_not_imported_list = set()
        poams_imported_list = []
        updated_systems_list = set()
        updated_poams_list = set()
        system_logs = []

        json_pattern = re.compile("^.*\.json$")
        xlsx_pattern = re.compile("^.*\.xlsx$")
        csv_pattern = re.compile("^.*\.csv$")
        
        file_import_record = ImportRecord.objects.create(name=file_name)
    
        if(json_pattern.match(file_name)):
            #JSON FILE
            is_ssp = json.loads(file_content).get("system-security-plan", None)
            if(is_ssp):
                new_system_name = is_ssp.get("system-characteristics").get("system-name")
                new_system_description = is_ssp.get("system-characteristics").get("description")
                new_system_msg = f"System {new_system_name} imported by upload"
                new_project, new_system = __create_new_system(request, new_system_name, new_system_description, new_system_msg)
                # messages.add_message(request, messages.INFO, new_system_msg)
                
                # import  all the other system characteristics
                security_sensitivity_level = is_ssp.get("system-characteristics").get("security-sensitivity-level")
                info_types = is_ssp.get("system-characteristics").get("system-information")
                security_impact_level = is_ssp.get("system-characteristics").get("security-impact-level")
                status = is_ssp.get("system-characteristics").get("status")
                system_authorization_boundary = is_ssp.get("system-characteristics").get("authorization-boundary")
                new_system.set_security_sensitivity_level(security_sensitivity_level)
                
                
                # import all users
                system_users = is_ssp.get("system-implementation").get("users")

                # import all of the components, check if it already exists, if it doesn't then create
                new_components_list = is_ssp.get("system-implementation").get("components")
                for comp in new_components_list:
                    # for each component in the imported components, check if it exists
                        # if it does, continue 
                        # else -> create

                        {
                            'uuid': '37107aa8-8cea-4a49-9222-cd157f075690', 
                            'title': 'Central Log Server Security Requirements Guide', 
                            'description': 'Central Log Server Security Requirements Guide based on DOD SRG/STIG.', 
                            'status': {'state': 'operational'}, 
                            'type': 'software', 
                            'responsible-roles': 
                                [{
                                    'role-id': 'asset-owner', 
                                    'party-uuids': ['7e909baa-9c1f-4d36-9992-6830ea461b7f']}
                                ], 
                            'props': []
                        }
                        newComponent = Element.objects.get_or_create(
                            title = comp["title"],
                            description = comp["description"],
                            component_type = comp["type"],
                            component_state = comp["status"]["state"],
                            uuid = comp["uuid"],
                        )
                        # Add responsible-roles to component
                        # Assign parties to component
                        Party.objects.get_or_create(
                            comp["responsible-roles"][0]["party-uuids"],
                        )
                        print(comp)

                # import all of the inventory items
                # import all of the control implementations

                sys_successful = Project.objects.filter(id=new_project.id).exists()

                if(sys_successful):
                    systems_imported_list.add(new_system)
                else: 
                    systems_not_imported_list.add({'name': new_system_name})

        if(xlsx_pattern.match(file_name)):
            content = ast.literal_eval(file_content)
            firstSheet = next(iter(content))
            sheet = content[firstSheet]
            
            for index, system in enumerate(sheet):
                # Assigning each column to appropriate variable
                # check if value in column is the correct format
                new_system_info = {}
                faulty_system = {}
                sys_name = ''
                sys_log = []
                list_of_system_properties = ["CSAM ID", "Org", "Sub Org", "System Name", "Acronym", "System Category", "System Operational Status", "System Type", "Contractor System", "Financial System", "FISMA Reportable", "Critical Infrastructure", "Mission Critical", "Investment Name", "Portfolio", "UII Code", "Investment Name"]
                list_of_poam_attributes = ["POAM Sequence", "Days Since Creation", "Scheduled Completion Date", "Planned Start Date", "Actual Start Date", "Planned Finish Date", "Actual Finish Date", "Weakness", "Cost", "Control Risk Severity", "User Identified Criticality", "Severity", "Workflow Status", "Workflow Status Date", "Days Until Auto-Approved", "Exclude From OMB", "Accepted Risk", "Assigned To", "Phone", "Email", "Assigned Date", "Delay Reason", "CSFFunction", "CSFCategory","CSFSubCategory", "Number Milestones", "Number Artifacts", "RBD Approval Date", "Deficiency Category", "Source of Finding", "Percent Complete", "Date % Complete Last Updated", "Delay Justification", "Monthly Status", "Comments"]
                wasSystemCreated = False
                hasSystemUpdated = False
                hasPoamUpdated = False
                hasPoamStmtUpdated = False

                if 'System Name' in system:
                    sys_name = system['System Name']
                    faulty_system[sys_name] = []
                if 'System Name' not in system:
                    sys_name = f"System on Row {index + 1}"
                    faulty_system[sys_name] = ["Missing System Name"]
                # if 'System Description' not in system:
                #     faulty_system[sys_name].append("Missing System Description")
                
                for column in system:
                    validated_value = validate_field(column, system[column])
                    if not isinstance(validated_value, dict):
                        new_system_info[column] = validated_value
                    else:
                        faulty_system[sys_name].append(validated_value)

                # check to see system has name and a description to create   
                if('System Name' in new_system_info and not faulty_system[sys_name]):
                    sys = System.objects.filter(root_element__name=new_system_info["System Name"]).first()
                    if sys:
                        proj = Project.objects.get(system=sys)
                        sys_log.append("System already exists.")
                    else:
                        new_system_msg = f"System imported from uploaded XLSX file"
                        proj, sys = __create_new_system(request, new_system_info['System Name'], "", new_system_msg, file_import_record)
                        wasSystemCreated = True
                        sys_log.append("Creating a new System and Project")
                    
                    # Add System related properties to info
                    for key in new_system_info:
                        if key in list_of_system_properties:
                            if key in sys.info:
                                if sys.info[key] != new_system_info[key]:
                                    hasSystemUpdated = True
                                    sys.info[key] = new_system_info[key]
                            else:
                                hasSystemUpdated = True
                                sys.info[key] = new_system_info[key]
                    sys.save_without_historical_record()

                    # Check if POAM already exists within System
                    if 'POAM ID' in new_system_info and new_system_info['POAM ID']:
                        poamExists = Poam.objects.filter(poam_id=new_system_info["POAM ID"]).first()
                        if poamExists:
                            sys_log.append("POAM already exists.")
                            # POAM exists, update values as needed
                            poamStatement = Statement.objects.get(id=poamExists.statement.id)
                            if poamExists.weakness_name != new_system_info["POAM Title"]:
                                poamExists.weakness_name = new_system_info["POAM Title"]
                                hasPoamUpdated = True
                                sys_log.append(f"POAM Title has been updated to: {poamExists.weakness_name}")
                            if poamStatement.body != new_system_info["Detailed Weakness Description"]:
                                poamStatement.body = new_system_info["Detailed Weakness Description"]
                                hasPoamStmtUpdated = True
                                sys_log.append(f"POAM Body has been updated to: {poamStatement.body}")
                            if poamStatement.statement_type != StatementTypeEnum.POAM.name:
                                poamStatement.statement_type = StatementTypeEnum.POAM.name
                                hasPoamStmtUpdated = True
                                sys_log.append(f"POAM Type has been updated to: {poamStatement.statement_type}")
                            if poamStatement.consumer_element != sys.root_element:
                                poamStatement.consumer_element = sys.root_element
                                hasPoamStmtUpdated = True
                                sys_log.append(f"POAM Consumer_Element has been updated to: {poamStatement.consumer_element}")
                            if poamStatement.status != new_system_info["Status"]:
                                poamStatement.status = new_system_info["Status"]
                                hasPoamStmtUpdated = True
                                sys_log.append(f"POAM Status has been updated to: {poamStatement.status}")
                            if poamStatement.created != new_system_info["Create Date"]:
                                poamStatement.created = new_system_info["Create Date"]
                                hasPoamStmtUpdated = True
                                sys_log.append(f"POAM Created has been updated to: {poamStatement.created}")
                            if poamExists.controls != new_system_info["Controls"]:
                                poamExists.controls = new_system_info["Controls"]
                                hasPoamUpdated = True
                                sys_log.append(f"POAM Controls has been updated to: {poamExists.controls}")
                            if poamExists.scheduled_completion_date != new_system_info["Planned Finish Date"]:
                                poamExists.scheduled_completion_date = new_system_info["Planned Finish Date"]
                                hasPoamUpdated = True
                                sys_log.append(f"POAM Scheduled Completion Date has been updated to: {poamExists.scheduled_completion_date}")
                            if int(poamExists.milestones) != int(new_system_info["Number Milestones"]):
                                poamExists.milestones = int(new_system_info["Number Milestones"])
                                hasPoamUpdated = True
                                sys_log.append(f"POAM Milestones has been updated to: {poamExists.milestones}")
                            
                            for key in new_system_info:
                                if key in list_of_poam_attributes:
                                    if key in poamExists.extra:
                                        if poamExists.extra[key] != new_system_info[key]:
                                            hasPoamUpdated = True
                                            poamExists.extra[key] = system[key]
                                    else:
                                        hasPoamUpdated = True
                                        poamExists.extra[key] = system[key]
                                    
                            poamStatement.save_without_historical_record()
                            poamExists.save_without_historical_record()
                            # Check if POAM has been updated at all, if it has update the updated field to statement and document import_record with this file
                            if hasPoamUpdated or hasPoamStmtUpdated:
                                poamStatement.updated = datetime.now()
                                poamStatement.import_record.add(file_import_record)
                                poamExists.import_record.add(file_import_record)
                                sys_log.append(f"POAM Import Record has been updated to: {file_import_record.name}")
                                poamStatement.save()
                                poamExists.save()
                            if not hasPoamUpdated:
                                sys_log.append(f"POAM did not update.")
                        else:
                            # Create a new POAM statement object, then a POAM object
                            poam_statement = Statement.objects.create(
                                sid = None,
                                sid_class = None,
                                body = new_system_info["Detailed Weakness Description"],
                                statement_type = StatementTypeEnum.POAM.name,
                                consumer_element = sys.root_element,
                                status = new_system_info["Status"],
                                created = new_system_info["Create Date"],
                                updated = datetime.now()
                            )
                            poam_statement.import_record.add(file_import_record)
                            poam_statement.created = new_system_info["Create Date"]
                            poam_statement.save_without_historical_record()
                            
                            new_poam = Poam.objects.create(
                                statement = poam_statement,
                                poam_id = new_system_info["POAM ID"],
                                controls = new_system_info["Controls"],
                                weakness_name = new_system_info["POAM Title"],
                                # weakness_detection_source = 
                                # weakness_source_identifier = 
                                # remediation_plan = 
                                scheduled_completion_date = new_system_info["Planned Finish Date"],
                                milestones = new_system_info["Number Milestones"],
                                # milestone_changes =
                                # risk_rating_original =
                                # risk_rating_adjusted =
                                # poam_group =
                            )
                            for key in new_system_info:
                                if key in list_of_poam_attributes:
                                    if "date" in key.lower(): #Checking if we are going to assign a datetime object
                                        new_poam.extra[key] = system[key]
                                    else:
                                        new_poam.extra[key] = new_system_info[key]
                            # Check if added attributes are in the excel doc, if not, add them
                            for attribute in ["Delay Justification", "Monthly Status", "Comments"]:
                                if attribute not in new_poam.extra:
                                    new_poam.extra[attribute] = ""
                            new_poam.save_without_historical_record()
                            sys_log.append(f"POAM {new_poam.poam_id} was created")
                    
                    # Insert everything into System's info dictionary
                    for key in new_system_info:
                        if key == "CSAM ID" and sys.info['other_id'] != new_system_info[key]:
                                sys.info["other_id"] = new_system_info[key]
                                sys.info["CSAM ID"] = new_system_info[key]
                                hasSystemUpdated = True
                                sys_log.append("Setting CSAM ID to System info as other_id")

                        # Check if key is already in system info
                        if key in sys.info:
                            # Check datetime objects
                            if "date" in key.lower():
                                try:
                                    temp_dt = new_system_info[key].strftime("%m/%d/%Y")
                                    if sys.info[key] != temp_dt:
                                        sys.info[key] = new_system_info[key].strftime('%m/%d/%Y')
                                        hasSystemUpdated = True
                                        sys_log.append(f"Updated {key} = {new_system_info[key].strftime('%m/%d/%Y')} into System Info Log")
                                except:
                                    print("EXCEPTION at key: ", key)
                            else:
                                sys.info[key] = new_system_info[key]
                                hasSystemUpdated = True
                                sys_log.append(f"Updated {key} = {new_system_info[key]} into System Info Log")
                        # Adding new key into system info
                        if "date" in key.lower() and key not in sys.info and new_system_info[key]:
                            if new_system_info[key]:
                                sys.info[key] = new_system_info[key].strftime('%m/%d/%Y')
                                sys_log.append(f"Added {key} = {new_system_info[key].strftime('%m/%d/%Y')} into System Info Log")
                            else:
                                sys.info[key] = ""
                                sys_log.append(f"Added {key} = '' into System Info Log")
                    sys.save_without_historical_record()
                    # Assigning properties as dictated in the excel document to the system
                    
                    # Setting system information 
                    # Check if uuid was assigned, if it was, check if valid UUID, if true => set, else throw as error and return as not implemented
                    if 'System UUID' in new_system_info and new_system_info['System UUID']:
                        sys.root_element.uuid = new_system_info['System UUID']
                        hasSystemUpdated = True
                        sys_log.append(f"Setting System UUID has been updated to: {sys.root_element.uuid}")
                    if 'System Private' in new_system_info and new_system_info['System Private']:
                        sys.root_element.private = new_system_info['System Private']
                        hasSystemUpdated = True
                        sys_log.append(f"Private has been updated to: {sys.root_element.private}")
                    if 'System Require Approval' in new_system_info and new_system_info['System Require Approval']:
                        sys.root_element.require_approval = new_system_info['System Require Approval']
                        hasSystemUpdated = True
                        sys_log.append(f"System Require Approval has been updated to: {sys.root_element.require_approval}")
                    if 'System Created' in new_system_info and new_system_info['System Created']:
                        sys.root_element.created = new_system_info['System Created']
                        hasSystemUpdated = True
                        sys_log.append(f"System Created has been updated to: {sys.root_element.created}")
                    if 'System Updated' in new_system_info and new_system_info['System Updated']:
                        sys.root_element.updated = new_system_info['System Updated']
                        hasSystemUpdated = True
                        sys_log.append(f"System Private has been updated to: {sys.root_element.updated}")
                    if 'System OSCAL Version' in new_system_info and new_system_info['System OSCAL Version']:
                        sys.root_element.oscal_version = new_system_info['System OSCAL Version']
                        hasSystemUpdated = True
                        sys_log.append(f"System Private has been updated to: {sys.root_element.oscal_version}")
                    if 'System Component Type' in new_system_info and new_system_info['System Component Type']:
                        sys.root_element.component_type = new_system_info['System Component Type']
                        hasSystemUpdated = True
                        sys_log.append(f"System Private has been updated to: {sys.root_element.component_type}")
                    if 'System Component State' in new_system_info and new_system_info['System Component State']:
                        sys.root_element.component_state = new_system_info['System Component State']
                        hasSystemUpdated = True
                        sys_log.append(f"System Private has been updated to: {sys.root_element.component_state}")

                    # if System has been updated at all, document file as it's import record
                    if wasSystemCreated:
                        sys.import_record.add(file_import_record)
                        sys_log.append(f"System created using this import_record: {file_import_record.name}")
                        sys.save_without_historical_record()
                    if not wasSystemCreated and hasSystemUpdated:
                        sys.import_record.add(file_import_record)
                        sys_log.append(f"System Import_Record has been updated to: {file_import_record.name}")
                        sys.save()
                    if not hasSystemUpdated and not hasPoamUpdated:
                        sys_log.append(f"Nothing has been updated on this system: {sys.root_element.name}")
                    sys_successful = Project.objects.filter(id=proj.id).exists()
                    if(sys_successful):
                        systems_imported_list.add(sys)
                    system_logs.append(
                        {
                            "name": sys.root_element.name, 
                            "log": sys_log
                        }
                    )
                if faulty_system[sys_name]:
                    systems_not_imported_list.add(faulty_system)
         
        if(csv_pattern.match(file_name)):
            # CSV FILE
            csv_systems = file_content.split('\n')
            csv_headers = csv_systems[0].split(',')
            csv_headers[-1] = csv_headers[-1].split('\r')
            csv_systems.pop(0)
            num_of_column = len(csv_headers)
            new_system_msg = f"System imported from uploaded CSV file"
            
            updated_csv_system = []
            systems_imported_list = []
            systems_not_imported_list = []

            for sys in csv_systems:
                if not sys:
                    continue
                sys = sys.split(',')
                sys.pop(-1)
                sys_dict = {}
                for x in range(0, num_of_column-1):
                    sys_dict[csv_headers[x]] = sys[x]
                updated_csv_system.append(sys_dict)
            
            for sys in updated_csv_system:
                # Check each system attribute to see if they follow correct format/information
                new_system_info = {}
                faulty_system = {}
                
                for key in sys.keys():
                    validated_value = validate_field(key, sys[key])
                    if isinstance(validated_value, dict):
                        faulty_system.append(validated_value)
                    else:
                        new_system_info[column] = validated_value
                
                if 'System Name' not in new_system_info:
                    faulty_system.append({"System Name": "Missing System Name"})
                # if 'System Description' not in new_system_info:
                #     faulty_system.append({"System Description": "Missing System Description"})

                if 'System Name' in new_system_info and 'System Description' in new_system_info and not faulty_system: 
                    new_project, new_system = __create_new_system(request, new_system_info['System Name'], new_system_info['System Description'], new_system_msg)
                    
                    if "System UUID" in new_system_info and new_system_info['System UUID']:
                        new_system.root_element.uuid = new_system_info['System UUID']
                    if "System Private" in new_system_info and new_system_info['System Private']:
                        new_system.root_element.private = new_system_info['System Private']
                    if "System Require Approval" in new_system_info and new_system_info['System Require Approval']:
                        new_system.root_element.require_approval = new_system_info['System Require Approval']
                    if "System Created" in new_system_info and new_system_info['System Created']:
                        new_system.root_element.created = new_system_info['System Created']
                    if "System Updated" in new_system_info and new_system_info['System Updated']:
                        new_system.root_element.updated = new_system_info['System Updated']
                    if "System OSCAL Version" in new_system_info and new_system_info['System OSCAL Version']:                        
                        new_system.root_element.oscal_version = new_system_info['System OSCAL Version']
                    if 'System Component Type' in new_system_info and new_system_info['System Component Type']:
                        new_system.root_element.oscal_version = new_system_info['System Component Type']
                    if 'System Component State' in new_system_info and new_system_info['System Component State']:
                        new_system.root_element.oscal_version = new_system_info['System Component State']

                    # Now we check if the system was successfully created
                    sys_successful = Project.objects.filter(id=new_project.id).exists()
                    if(sys_successful):
                        systems_imported_list.append(new_system)
                else:
                    systems_not_imported_list.append(faulty_system)

    context = {
        "import_uuid": str(file_import_record.uuid),
        "systems_imported_list": systems_imported_list,
        "systems_not_imported_list": systems_not_imported_list,
        "updated_systems_list": updated_systems_list,
        "updated_poams_list": updated_poams_list,
        "system_logs": system_logs,
    }
    return render(request, "systems/new_systems_imported.html", context)

def undo_import(object_type, import_record_uuid):
    import types
    type_system = type(System.objects.first())
    if object_type == None:
        pass
    else:
        for obj in type(object_type).objects.filter(import_record__uuid=import_record_uuid):
            if obj.history.all().count() == 1:
                # object was just created from the import, we have to delete now
                # If system also delete system element
                if type(object_type) == type_system:
                    element = Element.objects.get(id=obj.root_element.id)
                    element.delete()
                obj.delete()
            elif obj.history.all().count() > 1:
                prev_obj = obj.history.first().prev_record
                # object was updated, so we must revert to the previous state
                for key, value in obj.__dict__.items():
                    if not isinstance(value, types.FunctionType or types.MethodType) and not key.startswith('_'):
                        setattr(obj, key, getattr(prev_obj, key))
                obj.save()
            else:
                pass

def revert_system_import(request):
    """
    Reverting system import
    """

    # Grab import record uuid
    # cycle through all records of Project/System/Poam/etc that has an import record
    # revert to previous historical state
    if request.method == 'POST':
        data = request.POST
        import_uuid = data.get('import_uuid')
        # Cycle through all import elements
        # if element was already in the system and no change
            # wouldnt have import_record, but we also wouldn't need to revert anything
        # if there was an update on the element, then element would have an import record with file_name and uuid
        # if element was created, then we need to delete
            # how do we tell which element was created from the system import
            # has IMPORT_RECORD

        undo_import(Project.objects.first(), import_uuid)
        undo_import(System.objects.first(), import_uuid)
        undo_import(Statement.objects.first(), import_uuid)
        undo_import(Poam.objects.first(), import_uuid)

        return HttpResponseRedirect("/projects")

@login_required
def project_import(request, project_id):
    """
    Import an entire project's components and control content
    """
    project = Project.objects.get(id=project_id)
    system_id = project.system.id
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    system_root_element = system.root_element

    # Retrieve identified System
    if request.method == 'POST':
        project_data = request.POST['json_content']
        # Need to get or create the app source by the id of the given app source
        #module_name = json.loads(project_data).get('project').get('module').get('key')
        title = json.loads(project_data).get('project').get('title')
        system_root_element.name = title

        logger.info(
            event="project JSON import update",
            object={"object": "project", "id": project.id, "title": project.title},
            user={"id": request.user.id, "username": request.user.username}
        )
        messages.add_message(request, messages.INFO, f'Updated project with id : {project.id}.')

        #Import questionnaire data
        log_output = []
        try:
            from collections import OrderedDict
            data = json.loads(project_data, object_pairs_hook=OrderedDict)
        except Exception as e:
            log_output.append("There was an error reading the export file.")
        else:
            try:
                # Update project data.
                project.import_json(data, request.user, "imp", lambda x: log_output.append(x))
            except Exception as e:
                log_output.append(str(e))

        # Log output
        logger.info(
            event="project JSON import",
            object={"object": "project", "id": project.id, "title": project.title, "log_output": log_output},
            user={"id": request.user.id, "username": request.user.username}
        )
        # TODO: Flags to avoid import any part of the project json (e.g. components, poams, questionnaire)
        # Import components and their statements
        loaded_imported_jsondata = json.loads(project_data)
        if loaded_imported_jsondata.get('component-definitions') != None:
            # Load and get the components then dump
            comp_num = 0
            for k, val in enumerate(loaded_imported_jsondata.get('component-definitions')):
                oscal_component_json = json.dumps(loaded_imported_jsondata.get('component-definitions')[k])
                import_name = request.POST.get('import_name', '')
                import_record = ComponentImporter().import_components_as_json(import_name, oscal_component_json, request)
                if import_record != None:
                    comps = add_selected_components(system, import_record)
                    comp_num = comp_num + len(comps)
            messages.add_message(request, messages.INFO, f"Created {comp_num} components.")

            # Import Poams
        if loaded_imported_jsondata.get('poams') != None:
            # Load and get the poams then dump
            poam_num = 0
            for k, poam in enumerate(loaded_imported_jsondata.get('poams')):
                # Create a Poam for the system
                ## Statement linked to the poam
                poamsmt_data = poam.get('statement')
                poam_smt = Statement.objects.create(
                    sid=None,
                    sid_class=None,
                    pid=None,
                    body= poamsmt_data.get('body'),
                    remarks= poam.get('remarks'),
                    version= poam.get('version'),
                    statement_type="POAM",
                    status= poamsmt_data.get('status', "New"),
                    uuid= str(uuid4()),
                    consumer_element= system_root_element
                )
                # Create Poam with statement and imported data
                poam = Poam.objects.create(statement = poam_smt, controls= poam.get('controls'), milestones = poam.get('milestones'), poam_id = Poam.objects.order_by('-poam_id')[0] + 1 if poam.get('poam_id') == None else poam.get('poam_id'),
                                           remediation_plan = poam.get('remediation_plan'), risk_rating_adjusted = poam.get('risk_rating_adjusted'),
                                           risk_rating_original = poam.get('risk_rating_original'), scheduled_completion_date = poam.get('scheduled_completion_date'),
                                           weakness_detection_source = poam.get('weakness_detection_source'), weakness_name = poam.get('weakness_name'),
                                           weakness_source_identifier = poam.get('weakness_source_identifier'), poam_group = poam.get('poam_group'))
                poam_num += 1
                logger.info(
                    event="Poam import",
                    object={"object": "poam", "id": poam.poam_id, "controls": poam.poam_group},
                    user={"id": request.user.id, "username": request.user.username}
                )
            messages.add_message(request, messages.INFO, f"Created {poam_num} Poams.")

        return HttpResponseRedirect("/projects")

@login_required
def project_export(request, project_id):
    """
    Export an entire project's components and control content
    """
    # Of the project in the current system. pick one project to export
    project = Project.objects.get(id=project_id)
    system_id = project.system.id
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    system_root_element = system.root_element

    # Retrieve related selected controls and Poams if user has permission on system
    if request.user.has_perm('view_system', system):

        # Iterate through the elements associated with the system get all statements produced for each
        oscal_comps = []
        for element in system.producer_elements:
            # Implementation statement OSCAL JSON
            impl_smts = element.statements_produced.filter(consumer_element=system_root_element)
            component = OSCALComponentSerializer(element, impl_smts).as_json()
            oscal_comps.append(component)
        poams = []
        poam_smts = system_root_element.statements_consumed.filter(statement_type="POAM").order_by('id')
        for smt in poam_smts:
            poam = {
                'controls': smt.poam.controls,
                'milestones': smt.poam.milestones,
                "poam_id": smt.poam.poam_id,
                'remediation_plan': smt.poam.remediation_plan,
                'risk_rating_original': smt.poam.risk_rating_original,
                'risk_rating_adjusted': smt.poam.risk_rating_adjusted,
                'scheduled_completion_date': smt.poam.scheduled_completion_date,
                'weakness_detection_source': smt.poam.weakness_detection_source,
                'weakness_name': smt.poam.weakness_name,
                "weakness_source_identifier": smt.poam.weakness_source_identifier,
                'poam_group': smt.poam.poam_group,
                "statement": {
                    "body": smt.body,
                    "remarks": smt.remarks,
                    "version": smt.remarks,
                    "status": smt.status,
                    "uuid": str(smt.uuid),
                }
            }
            json_serializer(poam)
            # Add json version as an element in the poams list
            poams.append(json.dumps(poam))
    questionnaire_data = json.dumps(project.export_json(include_metadata=True, include_file_content=True))
    data = json.loads(questionnaire_data)
    data['component-definitions'] = [json.loads(oscal_comp) for oscal_comp in oscal_comps]
    data['poams'] = [json.loads(poam) for poam in poams]
    response = JsonResponse(data, json_dumps_params={"indent": 2})
    filename = project.title.replace(" ", "_") + "-" + datetime.now().strftime("%Y-%m-%d-%H-%M")
    response['Content-Disposition'] = f'attachment; filename="{quote(filename)}.json"'
    return response

# System OSCAL
@login_required
def system_profile_oscal_json(request, system_id):
    """
    Return an OSCAL profile for this system.
    TODO: for now, we return an empty response.
    """

    data = {}
    return JsonResponse(data)
    response['Content-Disposition'] = f'attachment; filename="oscal-profile.json"'
    return response

# System Summaries
@login_required
def get_system_info(request, system_id):
    """Returns info for system"""

    system = System.objects.get(id=system_id)
    system_info = system.info
    print(system_info)
    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_info['id'] = system.id
    system_info['name'] = system.root_element.name
    # system_info['impact'] = system.get_security_impact_level

    system_info_default = {
        "system_description": "~",
        "id": "~",
        "other_id": "~",
        "name": "~",
        "organization_name": "~",
        "aka": "~",
        "impact": "~",
        "status": "~",
        "type": "~",
        "created": "~",
        "hosting_facility": "~",
        "next_audit": "~",
        "next_scan": "~", #"05/01/22",
        "security_scan": "~",
        "pen_test": "~", #"Scheduled for 05/05/22",
        "config_scan": "~",
        "purpose": "~",
        "vuln_new_30days": "~",
        "vuln_new_rslvd_30days": "~",
        "vuln_90days": "~",
        "risk_score": "~",
        "score_1": "~",
        "score_2": "~",
        "score_3": "~",
        "score_4": "~",
        "score_5": "~",
    }

    # Set any missing defaults
    for key in system_info_default:
        if key not in system_info:
            system_info[key] = system_info_default[key]
    print(system_info)
    return system_info

@login_required
def get_system_events(request, system_id):
    """Returns Events for system"""

    system = System.objects.get(id=system_id)
    system_events = [ { "event_tag": e.event_type, "event_summary": e.description} for e in system.events.all()]

    # # Retrieve events from integrations
    # csam_system_id = system.info.get('csam_system_id', None)
    # system_events = [
    #     { "event_tag": "TEST", "event_summary": "Penetration test scheduled - Automated penetration test run by SOC"},
    #     { "event_tag": "SCAN", "event_summary": "Security scan scheduled - Automated weekly security scan will occur Sunday, June 9"},
    #     { "event_tag": "SYS", "event_summary": "ISSO appointed - Janice Avery (contracor) has been appointed as ISSO for System"}
    # ]
    return system_events

@login_required
def get_integrations_system_info(request, system_id):
    """Returns Integrations info for system"""

    system = System.objects.get(id=system_id)
    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_name = system.root_element.name

    # Retrieve info from integrations
    csam_system_id = system.info.get('csam_system_id', None)
    if csam_system_id is not None:
        from integrations.models import Endpoint
        endpoints = Endpoint.objects.filter(endpoint_path=f"/v1/systems/{csam_system_id}")
        if len(endpoints) == 1:
            # found cached csam data
            csam_data = endpoints[0].data
            # {'id': 111,
            #  'externalId': 'string',
            #  'name': 'System A',
            #  'description': 'This is a simple test system',
            #  'acronym': 'string',
            #  'organization': 'string',
            #  'subOrganization': 'string',
            #  'operationalStatus': 'string',
            #  'systemType': 'string',
            #  'financialSystem': 'string',
            #  'classification': 'string',
            #  'contractorSystem': True,
            #  'fismaReportable': True,
            #  'criticalInfrastructure': True,
            #  'missionCritical': True,
            #  'purpose': 'string',
            #  'ombExhibit': 'string',
            #  'uiiCode': 'string',
            #  'investmentName': 'string',
            #  'portfolio': 'string',
            #  'priorFyFunding': 0,
            #  'currentFyFunding': 0,
            #  'nextFyFunding': 0,
            #  'categorization': 'string',
            #  'fundingImportStatus': 'string'}
            purpose = f"{csam_data.get('purpose_short', 'Missing')}"
            organization_name = f"{csam_data.get('organization', 'Missing')} {csam_data.get('subOrganization', 'Missing')}"
            other_id = f"{csam_data.get('id', 'Missing')}"
            system_type = f"{csam_data.get('systemType', 'Missing')}"
            status = f"{csam_data.get('operationalStatus', 'Missing')}"
            impact = f"{csam_data.get('categorization', 'Missing')}"
        else:
            csam_data = {'id': 111,
             'externalId': 'string',
             'name': 'System A',
             'description': 'This is a simple test system',
             'acronym': 'string',
             'organization': 'string',
             'subOrganization': 'string',
             'operationalStatus': 'string',
             'systemType': 'string',
             'financialSystem': 'string',
             'classification': 'string',
             'contractorSystem': True,
             'fismaReportable': True,
             'criticalInfrastructure': True,
             'missionCritical': True,
             'purpose': 'string',
             'ombExhibit': 'string',
             'uiiCode': 'string',
             'investmentName': 'string',
             'portfolio': 'string',
             'priorFyFunding': 0,
             'currentFyFunding': 0,
             'nextFyFunding': 0,
             'categorization': 'string',
             'fundingImportStatus': 'string'}
            purpose = f"{csam_data.get('purpose_short', 'Missing')}"
            organization_name = f"{csam_data.get('organization', 'Missing')} {csam_data.get('subOrganization', 'Missing')}"
            other_id = f"{csam_data.get('id', 'Missing')}"
            system_type = f"{csam_data.get('systemType', 'Missing')}"
            status = f"{csam_data.get('operationalStatus', 'Missing')}"
            impact = f"{csam_data.get('categorization', 'Missing')}"
    else:
        # Couldn't find CSAM system
        # Sample systems from DHS SORNs
        dhs_sorns = [
            { "name": "DHS/ALL-037 E-Authentication Records System of Records", "purpose": """This system collects information in order to authenticate an individual's identity for the purpose of obtaining a credential to electronically access a DHS program or application. This system includes DHS programs or applications that use a third-party identity service provider to provide any of the following credential services: Registration, including identity proofing, issuance, authentication, authorization, and maintenance. This system collects information that allows DHS to track the use of programs and applications for system maintenance and troubleshooting. The system also enables DHS to allow an individual to reuse a credential received when applicable and available."""},
            { "name": "DHS/ALL-032 Official Passport Application and Maintenance Records", "purpose": """The purpose of this system is to collect and maintain a copy of an official passport application or maintenance record on DHS employees and former employees, including political appointees, civilian, and military personnel (and dependents and family members that accompany military members assigned outside the continental United States) assigned or detailed to the Department, individuals who are formally or informally associated with the Department, including advisory committee members, employees of other agencies and departments in the federal government, and other individuals in the private and public sector who are on official business with the Department, who in their official capacity, are applying for an official passport or updating their official passport records where a copy is maintained by the Department."""},
            { "name": "DHS/ALL-034 Emergency Care Medical Records", "purpose": """The purpose of this system is to support MQM oversight to ensure consistent quality medical care and standardize the documentation of care rendered by DHS EMS medical care providers in diverse environments.""" },
            { "name": "DHS/ALL-035 Common Entity Index Prototype (CEI Prototype)", "purpose": """The purpose of this prototype is to determine the feasibility of establishing a centralized index of select biographic information that will allow DHS to provide a consolidated and correlated identity, thereby facilitating and improving DHS's ability to carry out its national security, homeland security, law enforcement, and benefits missions.""" },
            { "name": "DHS/ALL-042 Personnel Networking and Collaboration System of Records.", "purpose": """The purpose of this system is to permit DHS's collection of biographical and professional information of current DHS employees, contractors, and grantees to facilitate connections and collaboration among individuals supporting the Department's mission; aid in the identification of individuals within an organization; and to ensure efficient collaboration within the Department.""" },
            { "name": "DHS/ALL-044 eRulemaking", "purpose": """The purpose of this system is to permit members of the public to review and comment on DHS rulemakings and notices. DHS will use any submitted contact information to seek clarification of a comment, respond to a comment when warranted, and for such other needs as may be associated with the rule making or notice process.""" },
            { "name": "DHS/FEMA-006 Citizen Corps Program", "url": "http://www.gpo.gov/fdsys/pkg/FR-2013-07-22/html/2013-17456.htm", "purpose": "The purpose of this system is to allow state, local, tribal, and territorial communities to setup and register Citizen Corps Councils and CERT programs. Also, this system provides a way for individuals to locate and contact Councils, CERTs, and other Citizen Corps partners for more information regarding volunteer programs and opportunities nation-wide. Additionally, this system uses surveys to assess and enhance communities' preparedness and to improve the effectiveness of the Citizen Corps Program."}
            # { "name": "", "purpose": """ """ },
        ]

        random.seed(system_id)
        other_id = random.randint(1,2000)
        impact = random.choice(["Low Impact", "Moderate Impact", "Moderate Impact", "Moderate Impact", "Moderate Impact", "High Impact"])
        status = random.choice(["Operational", "Operational", "Operational", "Operational", "Operational", "Operational", "Operational", "Operational", "Under Development", "Planned"])
        system_from_sorn = random.choice(dhs_sorns)

        # Fix purpose
        if "purpose" not in system_from_sorn:
            purpose = "Missing"
        elif len(system_from_sorn['purpose']) > 750:
            purpose = system_from_sorn['purpose'][0:500]
        else:
            purpose = system_from_sorn['purpose']

        # system_name = system_from_sorn['name'].strip(" ").strip(",")
        if re.search(r"DHS/[A-Za-z/&0-9-]+[0-9]{0,4} ", system_name):
            system_name = re.sub(r"DHS/[A-Za-z/&0-9-]+[0-9]{0,4} ", '', system_name)

        # Organization
        if re.search(r"DHS/([A-Za-z]{0,4})", system_from_sorn['name'].strip(" ").strip(",")):
            organization_name = "DHS " + re.search(r"DHS/([A-Za-z]{0,4})", system_from_sorn['name'].strip(" ").strip(",")).group(1).strip()
        else:
            organization_name = "DHS"

        system_type = random.choice(["General Support System", "Major Application", "Major Application", "Major Application", "Major Application", "Major Application", "Major Application", "Minor Application", ])

    # Acronym
    if "(" in system_name:
        acronym = re.search(r"\((.*)\)", system_name).group(1).strip()
        system_name = re.sub(r" \(.*\)", '', system_name)
    else:
        acronym = "".join([word[0].upper() for word in system_name.replace("(","").replace(")","").split(" ")])
        if len(acronym) > 5:
            acronym = acronym[0:3]
        if len(acronym) == 1:
            acronym = system_name
    aka = [acronym]
    if len(system_name.split(" ")) > 7:
        short_name = " ".join([word for word in system_name.split(" ")[0:2]]) + " System"
        aka.append(short_name)

    hosting_facility = random.choice(["DISC", "AWS", "AWS","AWS", "AWS", "DC-1", "AWS", "AWS", "Azure", "AWS", "AWS", "AWS", "DC-1",  ])
    # Dates
    from datetime import datetime, date, timedelta, timezone
    date1, date2 = datetime(2014, 6, 3, tzinfo=timezone.utc), datetime(2022, 2, 1, tzinfo=timezone.utc)
    dates_between = date2 - date1
    total_days = dates_between.days
    created = date1 + timedelta(days=random.randrange(total_days))
    next_audit = datetime.now().date() + timedelta(random.randint(40,600))
    next_scan = datetime.combine(datetime.now().date() + timedelta(random.randint(2,12)), datetime.min.time()).replace(tzinfo=timezone.utc)

    # datetime.combine(date.today(), datetime.min.time())
    # next_scan = next_scan.replace(tzinfo=timezone.utc)
    dates_between_created = created - date1
    pen_test = created + timedelta(days=random.randrange(dates_between_created.days))

    # Data for System
    system = {
        "id": system_id,
        "other_id": other_id,
        "name": system_name,
        "organization_name": organization_name,
        "aka": aka,
        "impact": impact,
        "status": status,
        "type": system_type,
        "created": created,
        "hosting_facility": hosting_facility,
        "next_audit": next_audit,
        "next_scan": next_scan, #"05/01/22",
        "security_scan": "Last known-04/20/22 @ 1:23pm",
        "pen_test": pen_test, #"Scheduled for 05/05/22",
        "config_scan": "Last known-01/17/20 @ 4:32pm",
        "purpose": purpose,
        "vuln_new_30days": random.randint(0,12),
        "vuln_new_rslvd_30days": random.randint(0,20),
        "vuln_90days": random.randint(0,15),
        "risk_score": random.randint(6,10) + round(random.random(), 1),
        "score_1": random.randint(6,10) + round(random.random(), 1),
        "score_2": random.randint(6,10) + round(random.random(), 1),
        "score_3": random.randint(6,10) + round(random.random(), 1),
        "score_4": random.randint(6,10) + round(random.random(), 1),
        "score_5": random.randint(6,10) + round(random.random(), 1),
    }
    return system

@login_required
def get_integrations_system_events(request, system_id):
    """Returns CSAM info for system"""

    system = System.objects.get(id=system_id)
    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]
    system_name = system.root_element.name

    # Retrieve events from integrations
    csam_system_id = system.info.get('csam_system_id', None)
    system_events = [
        { "event_tag": "TEST", "event_summary": "Penetration test scheduled - Automated penetration test run by SOC"},
        { "event_tag": "SCAN", "event_summary": "Security scan scheduled - Automated weekly security scan will occur Sunday, June 9"},
        { "event_tag": "SYS", "event_summary": "ISSO appointed - Janice Avery (contracor) has been appointed as ISSO for System"}
    ]
    return system_events

def __create_new_system(request, new_system_name, new_system_description="Missing system description", new_system_msg="Missing new system message", import_record=None):
    """Create and return a new system and project from new system name"""

    # We'll create a new system by leveraging existing to code to
    # start a project from template library. We'll get our default
    # project template from library, create the project which will
    # also give us a new system. Then we will update name of system.
    # Finally, we'll return the information to the caller.

    # Get a project template - Use Speedy SSP as default
    # TODO: What if Speedy SSP has been deleted?
    source_slug = "govready-q-files-startpack"
    app_name = "speedyssp"
    # can user start the app?
    # Is this a module the user has access to? The app store
    # does some authz based on the organization.
    catalog = get_compliance_apps_catalog_for_user(request.user)
    for app_catalog_info in catalog:
        if app_catalog_info["key"] == source_slug + "/" + app_name:
            # We found it.
            break
    else:
        raise Http404()
    # Start the most recent version of the app.
    appver = app_catalog_info["versions"][0]
    organization = Organization.objects.first()  # temporary
    default_folder_name = "Started Apps"
    folder = Folder.objects.filter(
        organization=organization,
        admin_users=request.user,
        title=default_folder_name,
    ).first()

    # Get a folder for project & system
    if not folder:
        folder = Folder.objects.create(organization=organization, title=default_folder_name)
        folder.admin_users.add(request.user)

    # Get default task and q - we can use None
    task = None
    q = None

    # Get portfolio for project & system
    if request.GET.get("portfolio"):
        portfolio = Portfolio.objects.get(id=request.GET.get("portfolio"))
    else:
        if not request.user.default_portfolio:
            request.user.create_default_portfolio_if_missing()
        portfolio = request.user.default_portfolio

    # Create new project by starting our app project template
    try:
        project = start_app(appver, organization, request.user, folder, task, q, portfolio)
    except ModuleDefinitionError as e:
        error = str(e)

    # Set up additional system information
    new_system = project.system 
    new_system.info = {
        "created_from_input": new_system_name,
        "system_description": new_system_description,
        "id": "~",
        "other_id": "~",
        "name": "~",
        "organization_name": "~",
        "aka": "~",
        "impact": "~",
        "status": "~",
        "type": "~",
        "created": "~",
        "hosting_facility": "~",
        "next_audit": "~",
        "next_scan": "~", #"05/01/22",
        "security_scan": "~",
        "pen_test": "~", #"Scheduled for 05/05/22",
        "config_scan": "~",
        "purpose": "~",
        "vuln_new_30days": "~",
        "vuln_new_rslvd_30days": "~",
        "vuln_90days": "~",
        "risk_score": "~",
        "score_1": "~",
        "score_2": "~",
        "score_3": "~",
        "score_4": "~",
        "score_5": "~",
    }
    if import_record is not None:
        project.import_record.add(import_record)
        new_system.import_record.add(import_record)
    new_system.add_event("SYS", new_system_msg)
    new_system.save_without_historical_record()
    project.save_without_historical_record()    

    # Change name of new system by changing name of new system's root element
    # Use the new_system_name for new name and numbered appendix to avoid name collisions
    nsre = new_system.root_element
    name_suffix = ""
    while Element.objects.filter(name=f"{new_system_name}{name_suffix}").exists():
        if name_suffix == "":
            name_suffix = 1
        else:
            name_suffix = str(int(name_suffix)+1)
    if name_suffix == "":
        nsre.name = new_system_name
    else:
        nsre.name = f"{new_system_name}{name_suffix}"
    nsre.save()

    # Update System Project title to new system name
    prt = project.root_task
    prt.title_override = nsre.name
    prt.save()

    # Log creation of new system
    logger.info(event=f"create_system_from_url url {new_system_name}",
            object={"object": "system", "id": new_system.id},
            user={"id": request.user.id, "username": request.user.username})

    return project, new_system

@login_required
def create_new_system(request):
    """Create a new system"""

    new_system_str = request.GET.get("s", None)
    if new_system_str is None:
        # Display page with various forms to create new system when no data received

        # Get default apps to display on page
        from siteapp.views import filter_app_catalog
        catalog, filter_description = filter_app_catalog(get_compliance_apps_catalog_for_user(request.user), request)
        # Group by category from catalog metadata.
        from collections import defaultdict
        catalog_by_category = defaultdict(lambda: {"title": None, "apps": []})
        for app in catalog:
            source_slug, _ = app["key"].split('/')
            app['source_slug'] = source_slug
            # print(f"1 keys: {app.keys()}")
            # print(f"2 key, title: {app['key']}, {app['title']}")
            for category in app["categories"]:
                catalog_by_category[category]["title"] = (category or "Uncategorized")
                # Only get default apps
                # TODO: Refactor this code
                organization = Organization.objects.first()  # temporary
                # if app['title'] in ['Blank Project', 'Speedy SSP', 'General IT System ATO for 800-53 (low)']:
                if app['title'] in organization.extra.get('default_appversion_name_list', []): # temporary
                    catalog_by_category[category]["apps"].append(app)

        # Sort categories by title and discard keys.
        catalog_by_category = sorted(catalog_by_category.values(), key=lambda category: (
            category["title"] != "Great starter apps",  # this category goes first
            category["title"].lower(),  # sort case insensitively
            category["title"],  # except if two categories differ only in case, sort case-sensitively
        ))

        context = {
            "apps": catalog_by_category,
            "import_system_form": ImportSystemForm(),
            "import_project_form": ImportProjectForm(),
        }
        return render(request, "systems/new_system_form.html", context)

    new_system_name = new_system_str
    new_system_msg = f"Created new System in GovReady based with name '{new_system_name}'."
    new_system_description = f"System created from name."

    # Create for new system based on URL
    if 'http://' in new_system_str or 'https://' in new_system_str:
        new_system_url = new_system_str
        new_system_msg = f"Created new System in GovReady based on URL '{new_system_url}'."
        new_system_name = new_system_url.replace("https://", "").replace("http://", "")
        new_system_description = f"System created from url."
        # Examine remote site for information
        # Handle error case of no URL

    # Create the new system and it's related project
    project, new_system = __create_new_system(request, new_system_name, new_system_msg)
    # messages.add_message(request, messages.INFO, new_system_msg)

    # Redirect to the new system/project.
    return HttpResponseRedirect(project.get_absolute_url())
    # return redirect(reverse('system_summary', args=[new_system.id]))

@login_required
def system_summary_1_aspen(request, system_id):
    """System Summary page experiment 1"""

    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    system_events = get_integrations_system_events(request, system_id)

    # Get all projects
    projects = system.projects.all()

    # Fix menu data for old vertical menu
    project.system.root_element.name = system_summary['name']
    project.root_task.title_override = system_summary['name']

    context = {
        "system": system_summary,
        #"project": project,
        "project": projects,
        "system_events": system_events,
        # "deployments": deployments,
        "display_urls": project_context(project)
    }
    return render(request, "systems/system_summary_1.html", context)

@login_required
def system_summary_aspen(request, system_id):
    """System Summary page experiment 2 live data"""

    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    # system_summary = get_integrations_system_info(request, system_id)
    system_summary = get_system_info(request, system_id)
    # system_events = get_integrations_system_events(request, system_id)
    system_events = get_system_events(request,system_id)

    # Get all projects
    projects = system.projects.all()

    # Fix menu data for old vertical menu
    project.system.root_element.name = system_summary['name']
    project.root_task.title_override = system_summary['name']

    context = {
        "system": system_summary,
        "project": project,
        "projects": projects,
        "system_events": system_events,
        # "deployments": deployments,
        "display_urls": project_context(project)
    }
    return render(request, "systems/system_summary_1.html", context)

@login_required
def system_integrations_aspen(request, system_id):
    """System Integrations page experiment 1"""

    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    system_events = get_integrations_system_events(request, system_id)

    # Fix menu data for old vertical menu
    project.system.root_element.name = system_summary['name']
    project.root_task.title_override = system_summary['name']

    # Get integrations
    general_integrations = []
    integrations = Integration.objects.all()
    for integration in integrations:
        general_integration = {
            "integration_name": integration.name,
            "integration_summary": integration.description,
            "integration_base_url": integration.config.get('base_url', None)
            }
        general_integrations.append(general_integration)

    # system_integrations = [
    #     { "integration_name": "CSAM", "integration_summary": "Provides access to system information stored in CSAM"},
    #     { "integration_name": "GITHUB", "integration_summary": "Includes recent developer changes to source code"},
    #     # { "event_tag": "SYS", "event_summary": "Isso appointed - Lorem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod..."}
    # ]

    context = {
        "system": system_summary,
        "project": project,
        "system_events": system_events,
        "system_integrations": general_integrations,
        # "deployments": deployments,
        "display_urls": project_context(project)
    }
    return render(request, "systems/system_integrations_aspen.html", context)

# System Deployments
@login_required
def system_deployments_aspen(request, system_id):
    """List deployments for a system"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    system_events = get_integrations_system_events(request, system_id)

    # Retrieve list of deployments for the system
    deployments = system.deployments.all().order_by(Lower('name'))
    # controls = system.root_element.controls.all()
    # poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('-updated')

    # Return the controls
    context = {
        "system": system_summary,
        "project": project,
        "deployments": deployments,
        "display_urls": project_context(project)
    }
    # return render(request, "systems/deployments_list.html", context)
    return render(request, "systems/deployments_list_aspen.html", context)

@login_required
def system_deployments(request, system_id):
    """List deployments for a system"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    system_summary = get_integrations_system_info(request, system_id)
    system_events = get_integrations_system_events(request, system_id)

    # Retrieve list of deployments for the system
    deployments = system.deployments.all().order_by(Lower('name'))
    # controls = system.root_element.controls.all()
    # poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('-updated')

    # Return the controls
    context = {
        "system": system_summary,
        "project": project,
        "deployments": deployments,
        "display_urls": project_context(project)
    }
    # return render(request, "systems/deployments_list.html", context)
    return render(request, "systems/deployments_list.html", context)

@login_required
def manage_system_deployment(request, system_id, deployment_id=None):
    """Form to create or edit system deployment"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        # User does not have permission to this system
        raise Http404

    di = get_object_or_404(Deployment, pk=deployment_id) if deployment_id else None
    if request.method == 'POST':
        form = DeploymentForm(request.POST, instance=di)
        if form.is_valid():
            form.save()
            deployment = form.instance
            # Create message to display to user
            if di:
                messages.add_message(request, messages.INFO, f'Deployment "{deployment.name}" edited.')
                logger.info(
                    event="edit_deployment",
                    object={"object": "deployment", "id": deployment.id, "name":deployment.name},
                    user={"id": request.user.id, "username": request.user.username}
                )
            else:
                messages.add_message(request, messages.INFO, f'Deployment "{deployment.name}" created.')
                logger.info(
                    event="create_deployment",
                    object={"object": "deployment", "id": deployment.id, "name":deployment.name},
                    user={"id": request.user.id, "username": request.user.username}
                )
            return redirect('system_deployments', system_id=system_id)
    else:
        if di is None:
            di = Deployment(system_id=system_id)
        form = DeploymentForm(instance=di)

    return render(request, 'systems/deployment_form.html', {
        "form": form,
        "deployment": di,
    })

@login_required
def deployment_history(request, system_id, deployment_id=None):
    """Returns the history for the given deployment"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        # User does not have permission to this system
        raise Http404

    from controls.models import Deployment
    full_dpt_history = None
    try:
        deployments = Deployment.objects.get(id=deployment_id)
        full_dpt_history = deployments.history.all()
    except Deployment.DoesNotExist:
        messages.add_message(request, messages.ERROR, f'The deployment id is not valid. Is this still a deployment in GovReady?')
    context = {
        "deployment": full_dpt_history,
        }
    return render(request, "systems/deployment_history.html", context)

@login_required
def system_deployment_inventory(request, system_id, deployment_id):
    """List system deployment inventory"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm('view_system', system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        # Retrieve list of deployments for the system
        deployments = system.deployments.all()
        deployment = get_object_or_404(Deployment, pk=deployment_id)
        # controls = system.root_element.controls.all()
        # poam_smts = system.root_element.statements_consumed.filter(statement_type="POAM").order_by('-updated')

        # inventory_items = [item for item in inventory_all if item["deployment_id"] == deployment_id]
        inventory_items = [item for item in deployment.inventory_items] if deployment.inventory_items != None else []

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "deployment": deployment,
            "inventory_items": inventory_items,
            # "controls": controls,
            # "poam_smts": poam_smts,
            # "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            # "enable_experimental_oscal": SystemSettings.enable_experimental_oscal,
        }
        return render(request, "systems/deployment_inventory.html", context)
    else:
        # User does not have permission to this system
        raise Http404

@login_required
def system_assessment_results_list(request, system_id=None):
    """List System Assessment Results for a system"""

    # Retrieve identified System
    if system_id:
        # Can user view this sytem?
        system = System.objects.get(id=system_id)
        if not request.user.has_perm('view_system', system):
            # User does not have permission to this system
            raise Http404

        system = System.objects.get(id=system_id)
        # Retrieve related selected controls if user has permission on system
        if not request.user.has_perm('view_system', system):
            # User does not have permission to this system
            raise Http404

        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        sars = system.system_assessment_result.all().order_by('created').reverse()

        # Retrieve user's API keys
        api_keys = request.user.get_api_keys()

        context = {
            "system": system,
            "project": project,
            "sars": sars,
            "api_key_ro": api_keys['ro'],
            "api_key_rw": api_keys['rw'],
            "api_key_wo": api_keys['wo'],
            "display_urls": project_context(project)
        }
        return render(request, "systems/sar_list.html", context)

@login_required
def view_system_assessment_result_summary(request, system_id, sar_id=None):
    """View Summary of System Assessment Results"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        # User does not have permission to this system
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]
    sar = get_object_or_404(SystemAssessmentResult, pk=sar_id) if sar_id else None

    # Get assessment targets results from wrapped SAR data
    sar_items = [item for item in sar.assessment_results['sar']] if sar.assessment_results != None else []

    # Get summary pass fail across all assessment results included collection
    # TODO: note high/low category
    summary = {}
    for param in ["pass", "fail", "other", "unknown", "error"]:
        summary[param] = sum(d[param] for d in sar_items if d and param in d)

    return render(request, 'systems/sar_summary.html', {
        "project": project,
        "sar": sar,
        "sar_items": sar_items,
        "assessment_results_json": json.dumps(sar.assessment_results, indent=4, sort_keys=True),
        "summary": summary,
        "display_urls": project_context(project)
    })

@login_required
def manage_system_assessment_result(request, system_id, sar_id=None):
    """Form to create or edit system assessment result"""

    # Can user view this system?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm('view_system', system):
        # User does not have permission to this system
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    sari = get_object_or_404(SystemAssessmentResult, pk=sar_id) if sar_id else None
    if request.method == 'POST':
        form = SystemAssessmentResultForm(request.POST, instance=sari)
        if form.is_valid():
            form.save()
            sar = form.instance
            # Create message to display to user
            if sari:
                messages.add_message(request, messages.INFO, f'System assessment result "{sar.name}" edited.')
                logger.info(
                    event="edit_system_assessment_result",
                    object={"object": "system_assessment_result", "id": sar.id, "name":sar.name},
                    user={"id": request.user.id, "username": request.user.username}
                )
            else:
                messages.add_message(request, messages.INFO, f'System assessment result "{sar.name}" created.')
                logger.info(
                    event="create_system_assessment_result",
                    object={"object": "system_assessment_result", "id": sar.id, "name":sar.name},
                    user={"id": request.user.id, "username": request.user.username}
                )
            return redirect('system_assessment_results_list', system_id=system_id)
    else:
        if sari is None:
            sari = SystemAssessmentResult(system_id=system_id)
        form = SystemAssessmentResultForm(instance=sari)
        # Filter deployments to current system
        form.fields["deployment"].queryset = Deployment.objects.filter(system__id=system_id)

    return render(request, 'systems/sar_form.html', {
        'form': form,
        'system_id': system_id,
        "display_urls": project_context(project)
    })

@login_required
def system_assessment_result_history(request, system_id, sar_id=None):
    """Returns the history for the given deployment system assessment result"""

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]

    # TODO check user permission to view
    full_sar_history = None
    try:
        sar = SystemAssessmentResult.objects.get(id=sar_id)
        full_sar_history = sar.history.all()
    except SystemAssessmentResult.DoesNotExist:
        messages.add_message(request, messages.ERROR, f'The system assessment result id is not valid. Is this still a system assessment result in GovReady?')
    context = {
        "deployment": full_sar_history,
        "display_urls": project_context(project)
    }
    return render(request, "systems/sar_history.html", context)

@login_required
def new_system_assessment_result_wazuh(request, system_id):
    """Returns a SAR info from Wazuh and adds to system"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # Validate data
        valid = True
        for param in ["wazuhhost_val", "user_val", "passwd_val", "agents_val"]:
            if param not in request.POST or request.POST[param] == "":
                valid = False
                messages.add_message(request, messages.WARNING, f"Please complete field {param.replace('_val','')}")
        if not valid:
            return HttpResponseRedirect(f"/systems/{system_id}/assessments")

        # Check user permissions
        system = System.objects.get(pk=system_id)
        if not request.user.has_perm('change_system', system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="delete_smt permission_denied",
                object={"object": "statement", "id": statement.id},  # todo - statement not defined anywhere - Greg
                user={"id": request.user.id, "username": request.user.username}
            )
            return HttpResponseForbidden(
                "Permission denied. {} does not have change privileges to system and/or project.".format(
                    request.user.username))

        from sec_srvc.wazuh import WazuhSecurityService
        wazuh_sec_svc = WazuhSecurityService()
        wazuh_sec_svc.setup(base_url=request.POST['wazuhhost_val'])

        authentication = wazuh_sec_svc.authenticate(request.POST['user_val'], request.POST['passwd_val'])

        if wazuh_sec_svc.is_authenticated:
            identifiers = request.POST['agents_val']
            extracted_data = wazuh_sec_svc.extract_data(authentication, identifiers)

            # TODO: Set deployment id
            deployment_uuid = None

            transformed_data = wazuh_sec_svc.transform_data(extracted_data, system_id, "Scan Title", "Scan description", deployment_uuid)
            loaded_data = wazuh_sec_svc.load_data(transformed_data)

            # Determine deployment_id from deployment_uuid
            # TODO: Make sure deployment is associated with system
            if deployment_uuid is None or deployment_uuid == "None":
                # When deployment is not defined, leave blank and attach SAR to system only
                deployment = None
                deployment_id = None
            else:
                deployment = Deployment.objects.get(uuid=deployment_uuid)
                deployment_id = deployment.id

            sar = SystemAssessmentResult(
                    name=transformed_data["metadata"]["title"],
                    description=transformed_data["metadata"]["description"],
                    system_id=transformed_data["metadata"]["system_id"],
                    deployment_id=deployment_id,
                    assessment_results=transformed_data
                    # assessment_results=json.loads(request.FILES.get('data').read().decode("utf8", "replace"))
                )
            sar.save()
            logger.info(
                event="create_system_assessment_result",
                object={"object": "system_assessment_result", "id": sar.id, "name":sar.name},
                user={"id": request.user.id, "username": request.user.username}
            )
            messages.add_message(request, messages.INFO, "Data from Wazuh retrieved and loaded")

        # Redirect
            return HttpResponseRedirect(f"/systems/{system_id}/assessments")

        else:
            # TODO: better handling of response code; 401, 301, etc.
            raise Exception(wazuh_sec_svc.error_msg['error'])